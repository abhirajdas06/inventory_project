import json
import traceback
from datetime import date
 
from django.contrib import messages
from django.http import JsonResponse
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db.models import OuterRef, Q, Subquery, Value, CharField
from django.db.models.functions import Coalesce, Concat
from django.contrib.postgres.aggregates import StringAgg
 
from apps.core.models import Product, SpareCategory, Brand
from apps.core.services import create_server_with_components, add_server_component
from apps.core.permissions import has_permission
from apps.inventory.models import InventoryFreezeRecord, InventoryTransaction
from apps.servers.models import Server, ServerComponent
 
 
# ── Create ────────────────────────────────────────────────
def _mark_latest_stock_in_user(product, user):
    if not product or not getattr(user, 'is_authenticated', False):
        return
    txn = InventoryTransaction.objects.filter(
        product=product,
        transaction_type='IN',
    ).order_by('-created_at').first()
    if txn:
        txn.performed_by = user
        txn.save(update_fields=['performed_by'])


def _exclude_out_or_frozen(qs):
    latest_freeze = InventoryFreezeRecord.objects.filter(
        product=OuterRef('product')
    ).order_by('-frozen_at', '-id')
    return qs.annotate(
        latest_freeze_status=Subquery(latest_freeze.values('status')[:1])
    ).exclude(latest_type='OUT').filter(
        Q(latest_freeze_status__isnull=True) | ~Q(latest_freeze_status='FROZEN')
    )


def add_server(request):
    if request.method == 'POST':
        try:
            server_data = {
                # server identity
                'machine_type':           request.POST.get('machine_type', ''),
                'machine_no':             request.POST.get('machine_no', ''),
                'service_tag':            request.POST.get('service_tag', ''),
                'model':                  request.POST.get('model', ''),
                'brand':                  request.POST.get('brand', ''),
                # cabinet fields  ← NEW
                'part_no':                request.POST.get('part_no', ''),
                'alt_part_no':            request.POST.get('alt_part_no', ''),
                'alt_serial_no':          request.POST.get('alt_serial_no', ''),
                'specs':                  request.POST.get('specs', ''),
                'qty':                    request.POST.get('qty', 1),
                'barcode':                request.POST.get('barcode', ''),
                # testing
                'testing_date':           request.POST.get('testing_date', ''),
                'tested_by':              request.POST.get('tested_by', ''),
                # status / location
                'status':                 request.POST.get('status', 'WORKING'),
                'location':               request.POST.get('location', ''),
                'reference_location':     request.POST.get('reference_location', ''),
                'parent_child_location':  request.POST.get('parent_child_location', ''),
                'remark':                 request.POST.get('remark', ''),
                'store_location':         request.POST.get('store_location', 'WH1'),
                'stock_status':           request.POST.get('stock_status', 'LIVE'),
            }
 
            comp_fields = [
                'spare_type', 'brand', 'model',
                'part_no', 'alt_part_no',
                'serial_no', 'alt_serial_no',
                'specs', 'barcode', 'qty',
                'working_status', 'location',
                'reference_location', 'parent_child_location', 'remark',
            ]
 
            lists      = {f: request.POST.getlist(f'comp_{f}[]') for f in comp_fields}
            total      = len(lists['serial_no'])
            components = []
 
            for i in range(total):
                row = {}
                for f in comp_fields:
                    try:
                        row[f] = lists[f][i]
                    except IndexError:
                        row[f] = ''
                components.append(row)
 
            server = create_server_with_components(server_data, components)
            _mark_latest_stock_in_user(server.product, request.user)
            for component in server.components.select_related('product'):
                _mark_latest_stock_in_user(component.product, request.user)
            messages.success(request, 'Server created successfully.')
            return redirect('server_list')
 
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, str(e))
 
    return render(request, 'servers/add_server.html', {
        'brands':           Brand.objects.all(),
        'spare_categories': SpareCategory.objects.all(),
    })
 
 
def _server_queryset():
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')

    return Server.objects.select_related(
        'product', 'brand'
    ).annotate(
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        latest_client   = Subquery(latest_txn.values('client_name')[:1]),
        latest_invoice  = Subquery(latest_txn.values('invoice_no')[:1]),
        latest_olf_dc   = Subquery(latest_txn.values('olf_dc_number')[:1]),
        latest_out_date = Subquery(latest_txn.values('stock_out_date')[:1]),
        component_search = StringAgg(
            Concat(
                Coalesce('components__spare_type', Value('')),
                Value(' '),
                Coalesce('components__part_no', Value('')),
                Value(' '),
                Coalesce('components__alt_part_no', Value('')),
                Value(' '),
                Coalesce('components__serial_no', Value('')),
                Value(' '),
                Coalesce('components__alt_serial_no', Value('')),
                Value(' '),
                Coalesce('components__barcode', Value('')),
                output_field=CharField(),
            ),
            delimiter=' ',
            distinct=True,
        ),
    ).order_by('-created_at')


# ── List ──────────────────────────────────────────────────
def server_list(request):
    servers = _exclude_out_or_frozen(_server_queryset())
 
    return render(request, 'servers/server_list.html', {
        'servers':          servers,
        'spare_categories': SpareCategory.objects.all(),   # ← this
        'brands':           Brand.objects.all(),
    })


def server_out_list(request):
    selected_status = request.GET.get('status', '').strip()
    all_sold_qs = _server_queryset().filter(latest_type='OUT')
    available_statuses = sorted(list(set(
        all_sold_qs.values_list('latest_status', flat=True)
    )))
    available_statuses = [s for s in available_statuses if s]

    servers = all_sold_qs
    if selected_status:
        servers = servers.filter(latest_status=selected_status)

    return render(request, 'servers/server_out_list.html', {
        'servers': servers,
        'available_statuses': available_statuses,
        'selected_status': selected_status,
        'can_stock_return': has_permission(request.user, 'stock_return'),
        'can_stock_out': has_permission(request.user, 'stock_out'),
    })


def server_faulty_list(request):
    servers = _server_queryset().filter(latest_status__in=('FAULTY', 'DAMAGED'))
    return render(request, 'servers/server_out_list.html', {
        'servers': servers,
        'available_statuses': [],
        'selected_status': 'FAULTY',
        'is_faulty': True,
        'can_stock_return': has_permission(request.user, 'stock_return'),
        'can_stock_out': has_permission(request.user, 'stock_out'),
    })


SERVER_EXPORT_HEADERS = [
    'Sr.No', 'Testing Date', 'Tested by/FE name', 'Machine Type', 'Machine no',
    'System Service Tag No', 'Model', 'Spares Type', 'Part No', 'Alt Part No',
    'Serial No', 'Alt Serial. No', 'Specs', 'Barcode No', 'QTY',
    'Working/Not working', 'Location', 'Reference Location', 'Parent-child Location',
    'Remark(describe exact issue)',
]


def _in_stock_components(server):
    """In-stock components of a server (excludes stocked-out / frozen)."""
    latest = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
    freeze = InventoryFreezeRecord.objects.filter(
        product=OuterRef('product')
    ).order_by('-frozen_at', '-id')
    return (
        server.components
        .select_related('product', 'product__category')
        .annotate(
            latest_type=Subquery(latest.values('transaction_type')[:1]),
            freeze_status=Subquery(freeze.values('status')[:1]),
        )
        .exclude(latest_type='OUT')
        .filter(Q(freeze_status__isnull=True) | ~Q(freeze_status='FROZEN'))
        .order_by('id')
    )


def export_servers(request, state='live'):
    """Grouped XLSX export: each server cabinet followed by its in-stock
    components, in the exact Server import-template layout."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    sold = state == 'sold'
    servers = _server_queryset().filter(latest_type='OUT') if sold else _exclude_out_or_frozen(_server_queryset())
    if sold:
        selected_status = request.GET.get('status', '').strip()
        if selected_status:
            servers = servers.filter(latest_status=selected_status)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Servers'
    ws.append(SERVER_EXPORT_HEADERS)
    header_fill = PatternFill(fill_type='solid', fgColor='DCEBFF')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    group = 0
    for server in servers:
        group += 1
        cabinet_serial = server.product.serial_no if server.product else ''
        # Cabinet row
        ws.append([
            group,
            server.testing_date or '',
            server.tested_by or '',
            server.machine_type or '',
            server.machine_no or '',
            server.service_tag or '',
            server.model or '',
            'CABINET',
            server.part_no or '',
            server.alt_part_no or '',
            cabinet_serial,
            server.alt_serial_no or '',
            server.specs or '',
            server.barcode or '',
            server.qty or 1,
            server.status or '',
            server.location or '',
            server.reference_location or '',
            '',  # cabinet has no parent
            server.remark or '',
        ])
        # In-stock component rows, grouped under the cabinet
        for c in _in_stock_components(server):
            ws.append([
                group,
                server.testing_date or '',
                server.tested_by or '',
                server.machine_type or '',
                server.machine_no or '',
                server.service_tag or '',
                server.model or '',
                c.spare_type or (c.product.category.name if c.product and c.product.category else ''),
                c.part_no or '',
                c.alt_part_no or '',
                c.serial_no or (c.product.serial_no if c.product else ''),
                c.alt_serial_no or '',
                c.specs or '',
                c.barcode or '',
                c.qty or 1,
                c.working_status or '',
                c.location or server.location or '',
                c.reference_location or '',
                cabinet_serial,  # parent-child → cabinet serial
                c.remark or '',
            ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="servers-{state}.xlsx"'
    return response
 
 
# ── Components API ────────────────────────────────────────
def server_components(request, server_id):
    """Returns all components with latest transaction info."""
    try:
        server = Server.objects.get(id=server_id)
    except Server.DoesNotExist:
        return JsonResponse({'components': []})
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    comps = server.components.select_related(
        'product', 'product__category'
    ).annotate(
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_client   = Subquery(latest_txn.values('client_name')[:1]),
        latest_invoice  = Subquery(latest_txn.values('invoice_no')[:1]),
        latest_olf_dc   = Subquery(latest_txn.values('olf_dc_number')[:1]),
        latest_out_date = Subquery(latest_txn.values('stock_out_date')[:1]),
    ).order_by('spare_type')
 
    data = []
    for c in comps:
        is_out = (c.latest_type == 'OUT')
        data.append({
            'id':          c.id,
            'product_id':  c.product.id,
            'spare_type':  c.spare_type or '',
            'part_no':     c.part_no    or '',
            'alt_part_no': c.alt_part_no or '',
            'serial_no':   c.serial_no  or c.product.serial_no,
            'alt_serial_no': c.alt_serial_no or '',
            'specs':       c.specs      or '',
            'barcode':     c.barcode    or '',
            'qty':         c.qty,
            'working_status': c.working_status or '',
            'location':    c.location   or '',
            'remark':      c.remark     or '',
            'store':       c.latest_location or '-',
            'status':      c.latest_status   or '-',
            'is_out':      is_out,
            'client':      c.latest_client   or '',
            'invoice':     c.latest_invoice  or '',
            'olf_dc':      c.latest_olf_dc   or '',
            'out_date':    str(c.latest_out_date) if c.latest_out_date else '',
        })
 
    return JsonResponse({'components': data, 'server_model': server.model})
 
 
# ── Add component to existing server ─────────────────────
def add_server_component_view(request):
    if request.method == 'POST':
        try:
            server_id = request.POST.get('server_id')
            server    = Server.objects.get(id=server_id)
 
            comp_fields = [
                'spare_type', 'brand', 'model',
                'part_no', 'alt_part_no',
                'serial_no', 'alt_serial_no',
                'specs', 'barcode', 'qty',
                'working_status', 'location',
                'reference_location', 'parent_child_location', 'remark',
            ]
 
            lists = {f: request.POST.getlist(f'comp_{f}[]') for f in comp_fields}
            total = len(lists['serial_no'])
            added = 0
 
            last_txn = InventoryTransaction.objects.filter(
                product=server.product
            ).order_by('-created_at').first()
 
            store_location = last_txn.store_location if last_txn else 'WH1'
            stock_status   = last_txn.stock_status   if last_txn else 'LIVE'
 
            for i in range(total):
                barcode = lists['barcode'][i].strip().upper() if i < len(lists['barcode']) else ''
                serial  = lists['serial_no'][i].strip().upper() if i < len(lists['serial_no']) else ''
 
                if not barcode and not serial:
                    continue
 
                # duplicate barcode check
                if barcode:
                    from apps.categories.models import Spare
                    if Spare.objects.filter(barcode=barcode).exists():
                        return JsonResponse({'success': False, 'error': f'Barcode already exists: {barcode}'})
 
                row = {}
                for f in comp_fields:
                    try:
                        row[f] = lists[f][i]
                    except IndexError:
                        row[f] = ''
 
                component = add_server_component(server, row, store_location, stock_status)
                _mark_latest_stock_in_user(component.product, request.user)
                added += 1
 
            return JsonResponse({'success': True, 'added': added})
 
        except Server.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Server not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e), 'trace': traceback.format_exc()})
 
    return JsonResponse({'success': False})
 
 
# ── Inline edit ───────────────────────────────────────────
def update_server_field(request):
    ALLOWED = ['remark']
 
    if request.method == 'POST':
        srv_id = request.POST.get('id')
        field  = request.POST.get('field')
        value  = request.POST.get('value')
 
        if field not in ALLOWED:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            srv = Server.objects.get(id=srv_id)
            setattr(srv, field, value)
            srv.save()
            return JsonResponse({'success': True})
        except Server.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
 
    return JsonResponse({'success': False})
 
