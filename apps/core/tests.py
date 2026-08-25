from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from django.core.management import call_command
from django.test import TestCase
from openpyxl import load_workbook

from apps.core.importers import HEADER_MAPS, IMPORT_LABELS, import_combined_stock_out_row
from apps.core.models import Product
from apps.inventory.models import InventoryTransaction


class StockOutImportTests(TestCase):
    def test_every_category_has_a_stock_out_template(self):
        for base in ('battery', 'card', 'cpu', 'harddisk', 'memory',
                     'railkit', 'sfp', 'networking_spare', 'controller', 'server'):
            key = f'{base}_stock_out'
            self.assertIn(key, HEADER_MAPS, f'Missing stock-out template for {base}')
            self.assertIn(key, IMPORT_LABELS)
            base_headers = list(HEADER_MAPS[base].keys())
            combined_headers = list(HEADER_MAPS[key].keys())
            # Stock In columns must come first, unchanged and in order.
            self.assertEqual(combined_headers[:len(base_headers)], base_headers)
            # Appended Stock Out columns at the end.
            self.assertEqual(
                combined_headers[len(base_headers):],
                ['Client Name', 'Invoice No', 'OLF / DC No', 'Stock Status', 'Stock Out Date'],
            )

    def test_combined_card_stock_out_creates_product_stock_in_and_stock_out(self):
        row = {
            'brand': 'INTEL',
            'oem': 'OEM',
            'brand_model_no': 'X520',
            'interface': 'FC',
            'part_no': 'PN-1',
            'alt_part_no': '',
            'serial_no': 'CARD-SO-1',
            'brand_serial_no_1': 'CARD-SO-1',
            'capacity': '',
            'port': '',
            'barcode': 'BC-SO-1',
            'location': 'Rack 1',
            'reference_location': '',
            'remark': '',
            'store_location': 'WH1',
            'stock_status': 'LIVE',
            # appended stock-out columns:
            'so_client_name': 'Acme',
            'so_invoice_no': 'INV-9',
            'so_olf_dc_number': 'OLF-9',
            'so_stock_status': 'SALE',
            'so_stock_out_date': '2026-06-26',
        }
        product = import_combined_stock_out_row('card', row)
        self.assertIsInstance(product, Product)

        txns = InventoryTransaction.objects.filter(product=product).order_by('created_at')
        types = list(txns.values_list('transaction_type', flat=True))
        self.assertIn('IN', types)
        self.assertEqual(types[-1], 'OUT')

        out = txns.filter(transaction_type='OUT').latest('created_at')
        self.assertEqual(out.stock_status, 'SALE')
        self.assertEqual(out.client_name, 'Acme')
        self.assertEqual(out.invoice_no, 'INV-9')
        self.assertEqual(out.olf_dc_number, 'OLF-9')
        self.assertEqual(str(out.stock_out_date), '2026-06-26')


class DailyExportTests(TestCase):
    def test_daily_inventory_export_command_writes_both_workbooks(self):
        with TemporaryDirectory() as tmpdir:
            call_command('export_daily_inventory', output_dir=tmpdir, verbosity=0)
            export_root = Path(tmpdir) / 'exports' / date.today().isoformat()
            live_file = export_root / f'inventory-live-{date.today().isoformat()}.xlsx'
            sold_file = export_root / f'inventory-stocked_out-{date.today().isoformat()}.xlsx'

            self.assertTrue(live_file.exists())
            self.assertTrue(sold_file.exists())

            live_wb = load_workbook(live_file, read_only=True)
            sold_wb = load_workbook(sold_file, read_only=True)
            self.assertIn('Servers', live_wb.sheetnames)
            self.assertIn('Servers', sold_wb.sheetnames)
            live_wb.close()
            sold_wb.close()

    def test_daily_export_email_sends_both_workbooks(self):
        from django.core import mail
        from django.test import override_settings

        with TemporaryDirectory() as tmpdir:
            with override_settings(
                EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                DAILY_REPORT_RECIPIENTS=['abhiraj@zacocomputer.com'],
            ):
                call_command('export_daily_inventory', '--email',
                             output_dir=tmpdir, verbosity=0)
                self.assertEqual(len(mail.outbox), 1)
                message = mail.outbox[0]
                self.assertEqual(message.to, ['abhiraj@zacocomputer.com'])
                names = sorted(attachment[0] for attachment in message.attachments)
                self.assertEqual(len(names), 2)
                self.assertTrue(all(n.endswith('.xlsx') for n in names))
