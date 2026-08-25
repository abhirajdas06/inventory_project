from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _clean_cell(value):
    """openpyxl cannot write timezone-aware datetimes — strip the tzinfo
    (converting to local time first) so exports don't blow up with real data."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _append_row(ws, values):
    ws.append([_clean_cell(v) for v in values])


def _safe_sheet_name(name):
    cleaned = ''.join(ch for ch in name if ch.isalnum() or ch in (' ', '_', '-')).strip()
    cleaned = cleaned[:31].strip()
    return cleaned or 'Sheet'


def _resolve_attr(obj, path):
    current = obj
    for part in path.split('.'):
        current = getattr(current, part, None)
        if current is None:
            return ''
    return current


def _style_header(ws):
    fill = PatternFill(fill_type='solid', fgColor='DCEBFF')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _write_category_sheet(ws, items, fields):
    ws.append([label for label, _ in fields] + [
        'Store', 'Status', 'Client', 'Invoice', 'OLF / DC No', 'Stock Out Date', 'Last Audit', 'Created On'
    ])
    _style_header(ws)
    for item in items:
        _append_row(ws, [
            _resolve_attr(item, path) for _, path in fields
        ] + [
            getattr(item, 'latest_location', '') or '',
            getattr(item, 'latest_status', '') or '',
            getattr(item, 'latest_client', '') or '',
            getattr(item, 'latest_invoice', '') or '',
            getattr(item, 'latest_olf_dc', '') or '',
            getattr(item, 'latest_out_date', '') or '',
            getattr(item, 'last_audit_date', '') or '',
            getattr(item, 'created_at', '') or '',
        ])


def _write_server_sheet(ws, sold=False):
    from apps.servers.views import SERVER_EXPORT_HEADERS, _exclude_out_or_frozen, _in_stock_components, _server_queryset

    servers = _server_queryset().filter(latest_type='OUT') if sold else _exclude_out_or_frozen(_server_queryset())
    ws.append(SERVER_EXPORT_HEADERS)
    _style_header(ws)

    group = 0
    for server in servers:
        group += 1
        cabinet_serial = server.product.serial_no if server.product else ''
        ws.append([
            group,
            server.testing_date or '',
            server.tested_by or '',
            server.machine_type or '',
            server.machine_no or '',
            server.service_tag or '',
            server.model or '',
            'CABINET',
            server.part_no or '',
            server.alt_part_no or '',
            cabinet_serial,
            server.alt_serial_no or '',
            server.specs or '',
            server.barcode or '',
            server.qty or 1,
            server.status or '',
            server.location or '',
            server.reference_location or '',
            '',
            server.remark or '',
        ])
        component_rows = server.components.select_related('product', 'product__category').all() if sold else _in_stock_components(server)
        for component in component_rows:
            ws.append([
                group,
                server.testing_date or '',
                server.tested_by or '',
                server.machine_type or '',
                server.machine_no or '',
                server.service_tag or '',
                server.model or '',
                getattr(component, 'spare_type', '') or (component.product.category.name if component.product and component.product.category else ''),
                component.part_no or '',
                component.alt_part_no or '',
                component.serial_no or (component.product.serial_no if component.product else ''),
                component.alt_serial_no or '',
                component.specs or '',
                component.barcode or '',
                getattr(component, 'qty', 1) or 1,
                getattr(component, 'working_status', '') or '',
                component.location or server.location or '',
                component.reference_location or '',
                cabinet_serial,
                component.remark or '',
            ])


def export_daily_inventory_snapshots(output_dir=None, export_date=None):
    from apps.categories.views import LIST_MODELS, _annotated_category_queryset

    export_date = export_date or timezone.localdate()
    export_root = Path(output_dir or settings.MEDIA_ROOT) / 'exports' / export_date.isoformat()
    export_root.mkdir(parents=True, exist_ok=True)

    outputs = {}
    for state in ('live', 'stocked_out'):
        sold = state == 'stocked_out'
        workbook = Workbook()
        workbook.remove(workbook.active)
        for kind, config in LIST_MODELS.items():
            qs = _annotated_category_queryset(config['model'], sold=sold)
            if sold:
                qs = qs.order_by('-latest_out_date', '-id')
            else:
                qs = qs.order_by('id')
            ws = workbook.create_sheet(_safe_sheet_name(config['label']))
            _write_category_sheet(ws, qs.iterator(chunk_size=1000), config['fields'])
        server_ws = workbook.create_sheet('Servers')
        _write_server_sheet(server_ws, sold=sold)

        file_name = f'inventory-{state}-{export_date.isoformat()}.xlsx'
        file_path = export_root / file_name
        workbook.save(file_path)
        outputs[state] = str(file_path)

    return outputs


def send_daily_inventory_email(recipients=None, output_dir=None, export_date=None):
    """Generate the daily snapshots and email them as Excel attachments.

    Recipients default to settings.DAILY_REPORT_RECIPIENTS. Returns a dict with
    the send result and the generated file paths.
    """
    from django.core.mail import EmailMessage

    export_date = export_date or timezone.localdate()
    outputs = export_daily_inventory_snapshots(output_dir=output_dir, export_date=export_date)

    recipients = recipients or list(getattr(settings, 'DAILY_REPORT_RECIPIENTS', []) or [])
    if not recipients:
        return {'sent': False, 'reason': 'no recipients configured', 'outputs': outputs}

    subject = f'Daily Inventory Report — {export_date.isoformat()}'
    body = (
        f'Attached are the automated inventory snapshots for {export_date.isoformat()}:\n\n'
        '  • inventory-live — everything currently in stock\n'
        '  • inventory-stocked_out — everything stocked out / sold\n\n'
        'Each workbook has one sheet per category plus a grouped Servers sheet.\n\n'
        'This is an automated message from InvenTrack.'
    )
    email = EmailMessage(
        subject=subject,
        body=body,
        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
        to=recipients,
    )
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    for path in outputs.values():
        p = Path(path)
        email.attach(p.name, p.read_bytes(), content_type)

    email.send(fail_silently=False)
    return {'sent': True, 'recipients': recipients, 'outputs': outputs}
