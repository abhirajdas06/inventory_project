from datetime import date, datetime

from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook

from apps.categories.models import ImportJob, NetworkingSpare
from apps.core.activity import log_activity, log_timeline
from apps.core.models import Brand, Product, SpareCategory
from apps.core.services import (
    add_server_component,
    create_product_and_card,
    create_product_and_cpu,
    create_product_and_harddisk,
    create_product_and_memory,
    create_product_and_railkit,
    create_product_and_sfp,
    create_product_and_spare,
    create_server_with_components,
)
from apps.inventory.models import InventoryTransaction
from apps.servers.models import Server


DEFAULT_STORE = 'WH1'
DEFAULT_STATUS = 'LIVE'


IMPORT_LABELS = {
    'battery': 'Battery',
    'card': 'Card',
    'controller': 'Controller',
    'cpu': 'CPU',
    'harddisk': 'Hard Disk',
    'memory': 'Memory',
    'networking_spare': 'Networking Spare',
    'railkit': 'Rail Kit',
    'server': 'Server',
    'sfp': 'SFP',
    'stock_out': 'Stock Out',
}


HEADER_MAPS = {
    'battery': {
        'Product': 'category',
        'Brand': 'brand',
        'Model': 'model',
        'Part No': 'part_no',
        'Alt Part No': 'alt_part_no',
        'Serial No': 'serial_no',
        'Alt Serial. no': 'alt_serial_no',
        'Specs': 'specs',
        'QTY': 'qty',
        'Barcode No': 'barcode',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remark(describe exact issue)': 'remark',
    },
    'card': {
        'Brand': 'brand',
        'OEM': 'oem',
        'Brand Model No': 'brand_model_no',
        'Interface(Sas/Sata/Fc/Scsi)': 'interface',
        'Part No': 'part_no',
        'Alt Part No': 'alt_part_no',
        'Brand Serial no 1': 'brand_serial_no_1',
        'Capacity': 'capacity',
        'Port': 'port',
        'Barcode': 'barcode',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remarks': 'remark',
    },
    'controller': {
        'Sr.No': 'import_group',
        'Brand': 'brand',
        'Product Category': 'spare_type',
        'Product': 'product_name',
        'Model': 'model',
        'Part No': 'part_no',
        'Alt Part No': 'alt_part_no',
        'Serial No': 'serial_no',
        'Alt Serial. no': 'alt_serial_no',
        'Specs': 'specs',
        'QTY': 'qty',
        'Barcode No': 'barcode',
        'LOCATION': 'location',
        'Reference Location': 'reference_location',
        'Parent-child Location': 'parent_child_location',
        'Remark(describe exact issue)': 'remark',
    },
    'cpu': {
        'Brand': 'brand',
        'Model': 'model',
        'Part No': 'part_no',
        'No of Cores': 'no_of_cores',
        'No of Threads': 'no_of_threads',
        'GHZ': 'ghz',
        'Frequency': 'frequency',
        'Cache': 'cache',
        'Barcode': 'barcode',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remark': 'remark',
    },
    'harddisk': {
        'Brand': 'brand',
        'OEM': 'oem',
        'Brand Model No': 'brand_model_no',
        'OEM Model No': 'oem_model_no',
        'Capacity': 'capacity',
        'RPM': 'rpm',
        'Interface(Sas/Sata/Fc)': 'interface',
        'Part No': 'part_no',
        'Alt Part no': 'alt_part_no',
        'Alt Fru No / Alt Part No (1)': 'alt_fru_1',
        'Alt Fru No / Alt Part No (2)': 'alt_fru_2',
        'Alt Fru No / Alt Part No (3)': 'alt_fru_3',
        'Retail Part No(Box Pack)': 'retail_part_no',
        'Spare Part No-Tray': 'spare_part_tray',
        'GPN Code': 'gpn_code',
        'Brand Serial no 1': 'brand_serial_no',
        'Oem Serial no': 'oem_serial_no',
        'Serial No': 'serial_no',
        'Firmware': 'firmware',
        'Size(2.5/3.5)': 'size',
        'Health': 'health',
        'GB/S': 'gb_s',
        'Barcode': 'barcode',
        'Tray -Barcode': 'tray_barcode',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remark': 'remark',
    },
    'memory': {
        'Brand': 'brand',
        'OEM': 'oem',
        'Model': 'model',
        'Part No 1': 'part_no_1',
        'Part No 2': 'part_no_2',
        'Part No 3': 'part_no_3',
        'Size': 'size',
        'Ram Type(Ecc Regd/Fb Dimm )': 'ram_type',
        'DDR Version(PC)': 'ddr_version',
        'Frequency': 'frequency',
        'Rank': 'rank',
        'Serial No': 'serial_no',
        'QTY': 'qty',
        'Barcode': 'barcode',
        'Location': 'location',
        'Reference location': 'reference_location',
        'Remark': 'remark',
    },
    'networking_spare': {
        'Product': 'category',
        'Brand': 'brand',
        'Part no': 'part_no',
        'Alt Part No': 'alt_part_no',
        'Serial No': 'serial_no',
        'Alt Serial No': 'alt_serial_no',
        'Specs': 'specs',
        'QTY': 'qty',
        'Barcode No': 'barcode',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remark (Describe exact Issue)': 'remark',
    },
    'railkit': {
        'Brand': 'brand',
        'Left/Right': 'side',
        'Part No': 'part_no',
        'Serial No': 'serial_no',
        'SPECS': 'specs',
        'Barcode': 'barcode',
        'Supported Model': 'supported_model',
        'QTY': 'qty',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remark(describe exact issue)': 'remark',
    },
    'sfp': {
        'Brand': 'brand',
        'Model': 'model',
        'Part no': 'part_no',
        'Descripition': 'description',
        'Fibre type (SMF/MMF)': 'fibre_type',
        'Data Rate': 'data_rate',
        'Serial no': 'serial_no',
        'Barcode': 'barcode',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Remark': 'remark',
    },
    'server': {
        'Testing Date': 'testing_date',
        'Tested by/FE name': 'tested_by',
        'Machine Type': 'machine_type',
        'Machine no': 'machine_no',
        'System Service Tag No': 'service_tag',
        'Model': 'server_model',
        'Spares Type': 'spare_type',
        'Part No': 'part_no',
        'Alt Part No': 'alt_part_no',
        'Serial No': 'serial_no',
        'Alt Serial. No': 'alt_serial_no',
        'Specs': 'specs',
        'Barcode No': 'barcode',
        'QTY': 'qty',
        'Working/Not working': 'working_status',
        'Location': 'location',
        'Reference Location': 'reference_location',
        'Parent-child Location': 'parent_child_location',
        'Remark(describe exact issue)': 'remark',
    },
    'stock_out': {
        'Serial No': 'serial_no',
        'Barcode No': 'barcode',
        'Client Name': 'client_name',
        'Invoice No': 'invoice_no',
        'OLF / DC No': 'olf_dc_number',
        'Stock Status': 'stock_status',
        'Stock Out Date': 'stock_out_date',
    },
}


CREATE_BY_KEY = {
    'battery': create_product_and_spare,
    'card': create_product_and_card,
    'cpu': create_product_and_cpu,
    'harddisk': create_product_and_harddisk,
    'memory': create_product_and_memory,
    'railkit': create_product_and_railkit,
    'sfp': create_product_and_sfp,
}


# ════════════════════════════════════════════════════════════
#  STOCK OUT IMPORT — per category
#
#  Every inventory category gets a "<key>_stock_out" import that
#  reuses the EXACT Stock In template columns and appends the
#  Stock Out columns at the end. One row → create product +
#  stock in + immediate stock out.
#
#  The appended targets are prefixed `so_` so they never collide
#  with the Stock In field names (especially stock_status).
# ════════════════════════════════════════════════════════════
STOCK_OUT_APPEND_COLUMNS = {
    'Client Name': 'so_client_name',
    'Invoice No': 'so_invoice_no',
    'OLF / DC No': 'so_olf_dc_number',
    'Stock Status': 'so_stock_status',
    'Stock Out Date': 'so_stock_out_date',
}

# Categories that support combined stock-out import (everything that
# creates inventory). New inventory models added here are picked up
# automatically.
STOCK_OUT_BASE_KEYS = [
    'battery', 'card', 'cpu', 'harddisk', 'memory',
    'railkit', 'sfp', 'networking_spare', 'controller', 'server',
]

for _base_key in STOCK_OUT_BASE_KEYS:
    _combined_key = f'{_base_key}_stock_out'
    HEADER_MAPS[_combined_key] = {
        **HEADER_MAPS[_base_key],
        **STOCK_OUT_APPEND_COLUMNS,
    }
    IMPORT_LABELS[_combined_key] = f'{IMPORT_LABELS[_base_key]} — Stock Out'


def clean_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_qty(value):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 1


def read_headers(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return [clean_value(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]


def count_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(clean_value(v) for v in row):
            total += 1
    return total


def start_import_job(model_key, upload, user):
    if model_key not in HEADER_MAPS:
        raise ValueError('Unknown import model')

    job = ImportJob.objects.create(
        model_key=model_key,
        upload=upload,
        created_by=user if getattr(user, 'is_authenticated', False) else None,
        status='PENDING',
    )

    headers = read_headers(job.upload.path)
    expected = set(HEADER_MAPS[model_key])
    if model_key == 'stock_out':
        expected = {'Serial No', 'Barcode No'}
    missing = sorted(expected - set(headers))
    if missing:
        job.status = 'FAILED'
        job.errors = [{'row': 1, 'error': f"Missing columns: {', '.join(missing)}"}]
        job.save(update_fields=['status', 'errors'])
        raise ValueError(job.errors[0]['error'])

    job.total_rows = count_rows(job.upload.path)
    job.save(update_fields=['total_rows'])
    return job


def process_import_chunk(job, chunk_size=100, user=None):
    if job.status == 'DONE':
        return job

    job.status = 'RUNNING'
    job.save(update_fields=['status'])

    wb = load_workbook(job.upload.path, read_only=True, data_only=True)
    ws = wb.active
    headers = [clean_value(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    field_map = HEADER_MAPS[job.model_key]
    processed_this_call = 0
    seen_data_rows = 0
    current_controller_serial = ''

    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(clean_value(v) for v in values):
            continue
        seen_data_rows += 1

        raw = dict(zip(headers, values))
        row = {target: clean_value(raw.get(source)) for source, target in field_map.items()}
        row['qty'] = normalize_qty(row.get('qty'))
        row['store_location'] = DEFAULT_STORE
        if job.model_key == 'stock_out':
            row['stock_status'] = row.get('stock_status') or 'SALE'
            row['stock_out_date'] = row.get('stock_out_date') or date.today().isoformat()
        else:
            row['stock_status'] = DEFAULT_STATUS

        if job.model_key == 'controller':
            spare_type = (row.get('spare_type') or '').strip().upper()
            serial = (row.get('serial_no') or '').strip().upper()
            parent_serial = (row.get('parent_child_location') or '').strip().upper()

            if spare_type == 'CONTROLLER' and serial:
                current_controller_serial = serial
            elif not parent_serial and current_controller_serial:
                row['parent_child_location'] = current_controller_serial

        if seen_data_rows <= job.processed_rows:
            continue
        if processed_this_call >= chunk_size:
            break

        try:
            with transaction.atomic():
                product = import_row(job.model_key, row, user=user)
                mark_latest_stock_user(product, user)
            job.success_count += 1
        except Exception as exc:
            job.error_count += 1
            errors = list(job.errors or [])
            errors.append({'row': seen_data_rows + 1, 'error': str(exc)})
            job.errors = errors[-100:]

        job.processed_rows += 1
        processed_this_call += 1

    if job.processed_rows >= job.total_rows:
        job.status = 'DONE'

    job.save(update_fields=[
        'processed_rows', 'success_count', 'error_count', 'errors', 'status', 'updated_at'
    ])
    return job


def mark_latest_stock_user(product, user):
    if not product or not getattr(user, 'is_authenticated', False):
        return
    txn = InventoryTransaction.objects.filter(product=product, transaction_type='IN').order_by('-created_at').first()
    if txn:
        txn.performed_by = user
        txn.save(update_fields=['performed_by'])
        log_activity(
            action='IMPORT_STOCK_IN',
            module='IMPORT',
            entity=product.name,
            entity_id=product.id,
            user=user,
            barcode='',
            warehouse=txn.store_location,
            location='',
            new_values={
                'transaction_id': txn.id,
                'model': product.model,
                'stock_status': txn.stock_status,
            },
            remarks='Imported stock in',
        )
        log_timeline(
            product=product,
            event_type='IMPORT',
            user=user,
            warehouse=txn.store_location,
            details={
                'transaction_type': 'IN',
                'stock_status': txn.stock_status,
            },
        )


def import_row(model_key, row, user=None):
    if model_key == 'networking_spare':
        return create_product_and_networking_spare(row, user=user)
    if model_key == 'controller':
        return import_controller_row(row, user=user)
    if model_key == 'server':
        return import_server_row(row, user=user)
    if model_key == 'stock_out':
        return import_stock_out_row(row, user=user)
    if model_key.endswith('_stock_out'):
        base_key = model_key[:-len('_stock_out')]
        return import_combined_stock_out_row(base_key, row, user=user)
    creator = CREATE_BY_KEY[model_key]
    return creator(row)


# Appended Stock Out columns (see STOCK_OUT_APPEND_COLUMNS) — a row is only
# stocked out when at least one of these carries a value; otherwise the row is
# stocked in only.
STOCK_OUT_DETAIL_FIELDS = (
    'so_client_name', 'so_invoice_no', 'so_olf_dc_number',
    'so_stock_status', 'so_stock_out_date',
)


def _row_has_stock_out_details(row):
    return any((row.get(field) or '').strip() for field in STOCK_OUT_DETAIL_FIELDS)


def import_combined_stock_out_row(base_key, row, user=None):
    """Stock in every row; stock out only rows that actually carry Stock Out
    details (Client / Invoice / OLF-DC / Stock Status / Stock Out Date).

    So in a 10-row file where only 2 rows have stock-out columns filled, those
    2 are stocked out and the other 8 are stocked in only.
    """
    if base_key not in HEADER_MAPS:
        raise ValueError(f'Unknown stock-out base model: {base_key}')

    try:
        product = _product_from_stock_out_row(row)
    except ValueError:
        product = import_row(base_key, row, user=user)

    # No stock-out details on this row → stock in only.
    if not _row_has_stock_out_details(row):
        return product

    so_row = {
        'serial_no': product.serial_no or '',
        'barcode': '',
        'client_name': row.get('so_client_name') or '',
        'invoice_no': row.get('so_invoice_no') or '',
        'olf_dc_number': row.get('so_olf_dc_number') or '',
        'stock_status': row.get('so_stock_status') or 'SALE',
        'stock_out_date': row.get('so_stock_out_date') or '',
    }
    _apply_stock_out(product, so_row, user=user)
    return product


def _product_from_stock_out_row(row):
    from apps.categories.models import Card, CPU, HardDisk, Memory, NetworkingSpare, RailKit, SFP, Spare
    from apps.servers.models import Server, ServerComponent

    serial = (row.get('serial_no') or '').strip().upper()
    barcode = (row.get('barcode') or '').strip().upper()

    if serial:
        product = Product.objects.filter(serial_no=serial).first()
        if product:
            return product

    if barcode:
        related_models = (
            Spare, Card, CPU, Memory, SFP, RailKit, HardDisk, NetworkingSpare, Server, ServerComponent,
        )
        for model in related_models:
            item = model.objects.filter(barcode__iexact=barcode).select_related('product').first()
            if item:
                return item.product

    raise ValueError('Product not found by Serial No or Barcode No')


def import_stock_out_row(row, user=None):
    product = _product_from_stock_out_row(row)
    return _apply_stock_out(product, row, user=user)


def _apply_stock_out(product, row, user=None):
    """Stock out an already-resolved product using a stock-out row dict.

    Shared by the standalone Stock Out import and the per-category combined
    Stock In + Stock Out import. Validates the stock status and date, cascades
    to server components, and writes activity + timeline entries."""
    from datetime import date

    from apps.inventory.views import (
        _create_stock_out_transaction,
        _latest_location,
        _latest_store_location,
        _stock_out_server_components,
        create_rental_record,
    )

    stock_status = (row.get('stock_status') or 'SALE').strip().upper()
    valid_statuses = {value for value, _ in InventoryTransaction.STOCK_STATUS}
    if stock_status not in valid_statuses:
        raise ValueError(f'Invalid stock status: {stock_status}')

    raw_date = row.get('stock_out_date')
    stock_out_date = date.today()
    if raw_date:
        try:
            stock_out_date = datetime.strptime(str(raw_date), '%Y-%m-%d').date()
        except ValueError:
            stock_out_date = date.today()

    store_location = _latest_store_location(product)
    client_name = row.get('client_name') or ''
    invoice_no = row.get('invoice_no') or ''
    olf_dc_number = row.get('olf_dc_number') or ''

    with transaction.atomic():
        _create_stock_out_transaction(
            product=product,
            store_location=store_location,
            stock_status=stock_status,
            stock_out_date=stock_out_date,
            client_name=client_name,
            invoice_no=invoice_no,
            user=user,
            olf_dc_number=olf_dc_number,
        )
        affected = _stock_out_server_components(
            product=product,
            store_location=store_location,
            stock_status=stock_status,
            stock_out_date=stock_out_date,
            client_name=client_name,
            invoice_no=invoice_no,
            user=user,
            olf_dc_number=olf_dc_number,
        )

        if stock_status == 'RENT':
            create_rental_record(
                product=product,
                store_location=store_location,
                stock_out_date=stock_out_date,
                client_name=client_name,
                invoice_no=invoice_no,
                olf_dc_number=olf_dc_number,
                user=user,
                remarks='Rented out (import)',
            )

    log_activity(
        action='IMPORT_STOCK_OUT',
        module='IMPORT',
        entity=product.name,
        entity_id=product.id,
        user=user,
        warehouse=store_location,
        location=_latest_location(product),
        new_values={
            'stock_status': stock_status,
            'stock_out_date': str(stock_out_date),
            'client_name': client_name,
            'invoice_no': invoice_no,
            'olf_dc_number': olf_dc_number,
            'components_stocked_out': affected,
        },
        remarks='Stock-out imported from Excel',
    )
    log_timeline(
        product=product,
        event_type='OUT',
        user=user,
        warehouse=store_location,
        location=_latest_location(product),
        remarks='Stock-out imported from Excel',
        details={'invoice_no': invoice_no, 'olf_dc_number': olf_dc_number, 'stock_status': stock_status},
    )
    return product


def create_product_and_networking_spare(data, user=None):
    category_name = (data.get('category') or 'NETWORKING SPARE').strip().upper()
    category, _ = SpareCategory.objects.get_or_create(name=category_name)
    brand_name = (data.get('brand') or '').strip().upper()
    brand = None
    if brand_name:
        brand, _ = Brand.objects.get_or_create(name=brand_name)

    product = Product.objects.create(
        category=category,
        serial_no=(data.get('serial_no') or '').strip().upper(),
        part_no=data.get('part_no'),
        brand=brand_name,
        model=category_name,
        name=f"{category_name} - {brand_name} - {data.get('part_no', '')}".strip(' -'),
    )
    NetworkingSpare.objects.create(
        product=product,
        brand=brand,
        part_no=data.get('part_no'),
        alt_part_no=data.get('alt_part_no'),
        alt_serial_no=data.get('alt_serial_no'),
        specs=data.get('specs'),
        qty=data.get('qty') or 1,
        barcode=data.get('barcode') or None,
        location=data.get('location'),
        reference_location=data.get('reference_location'),
        remark=data.get('remark'),
    )
    from apps.core.services import _create_stock_in_transaction
    _create_stock_in_transaction(
        product=product,
        store_location=data.get('store_location') or DEFAULT_STORE,
        stock_status=data.get('stock_status') or DEFAULT_STATUS,
        user=user,
        remarks='Networking spare import stock-in',
    )
    return product


def import_controller_row(row, user=None):
    from apps.categories.models import Controller, Spare
    from apps.core.services import create_controller_with_components

    spare_type = (row.get('spare_type') or row.get('product_category') or '').strip().upper()
    product_name = (row.get('product_name') or row.get('product_type') or '').strip()
    serial = (row.get('serial_no') or '').strip().upper()
    barcode = (row.get('barcode') or '').strip().upper()
    parent_serial = (row.get('parent_child_location') or '').strip().upper()

    if spare_type == 'CONTROLLER' or product_name.upper() == 'CABINET':
        existing = Controller.objects.filter(product__serial_no=serial).select_related('product').first()
        if existing:
            return existing.product

        controller_data = dict(row)
        controller_data['model'] = row.get('model') or 'CONTROLLER'
        controller_data['serial_no'] = serial
        controller_data['spare_type'] = 'CONTROLLER'
        controller_data['product_name'] = product_name or 'CONTROLLER'
        controller = create_controller_with_components(controller_data, [])
        mark_latest_stock_user(controller.product, user)
        return controller.product

    if not parent_serial:
        raise ValueError('Parent-child Location must contain parent controller serial for controller components')

    controller = Controller.objects.filter(
        Q(product__serial_no=parent_serial) | Q(alt_serial_no=parent_serial)
    ).select_related('product').first()
    if not controller:
        raise ValueError(f'Parent controller not found for parent serial: {parent_serial}')

    existing_product = None
    if serial:
        existing_product = Product.objects.filter(serial_no=serial).first()
    elif barcode:
        existing_spare = Spare.objects.select_related('product').filter(barcode__iexact=barcode).first()
        if existing_spare:
            existing_product = existing_spare.product

    if existing_product:
        existing_component = Spare.objects.filter(
            controller=controller,
            product=existing_product,
        ).first()
        if existing_component:
            return existing_product

        orphan_controller = getattr(existing_product, 'controller', None)
        if orphan_controller and orphan_controller.pk != controller.pk and orphan_controller.components.count() == 0:
            orphan_controller.delete()

        brand_name = (row.get('brand') or '').strip().upper()
        brand = None
        if brand_name:
            brand, _ = Brand.objects.get_or_create(name=brand_name)

        category_name = spare_type or 'SPARE'
        category, _ = SpareCategory.objects.get_or_create(name=category_name)
        existing_product.category = category
        existing_product.part_no = row.get('part_no')
        existing_product.brand = brand_name
        existing_product.model = row.get('model')
        existing_product.name = (
            product_name
            or f"CONTROLLER - {category_name} - {row.get('model', '')} - {brand_name}".strip(' -')
        )
        existing_product.save(update_fields=['category', 'part_no', 'brand', 'model', 'name'])

        spare_record = getattr(existing_product, 'spare', None)
        if spare_record:
            spare_record.controller = controller
            spare_record.brand = brand
            spare_record.model = row.get('model')
            spare_record.part_no = row.get('part_no')
            spare_record.alt_part_no = row.get('alt_part_no')
            spare_record.alt_serial_no = row.get('alt_serial_no')
            spare_record.specs = row.get('specs')
            spare_record.qty = row.get('qty') or 1
            spare_record.barcode = barcode or None
            spare_record.location = row.get('location') or controller.location
            spare_record.reference_location = row.get('reference_location') or controller.reference_location
            spare_record.remark = row.get('remark')
            spare_record.save()
        else:
            Spare.objects.create(
                product=existing_product,
                controller=controller,
                brand=brand,
                model=row.get('model'),
                part_no=row.get('part_no'),
                alt_part_no=row.get('alt_part_no'),
                alt_serial_no=row.get('alt_serial_no'),
                specs=row.get('specs'),
                qty=row.get('qty') or 1,
                barcode=barcode or None,
                location=row.get('location') or controller.location,
                reference_location=row.get('reference_location') or controller.reference_location,
                remark=row.get('remark'),
            )
        return existing_product

    category_name = spare_type or 'SPARE'
    category, _ = SpareCategory.objects.get_or_create(name=category_name)

    brand_name = (row.get('brand') or '').strip().upper()
    brand = None
    if brand_name:
        brand, _ = Brand.objects.get_or_create(name=brand_name)

    product = Product.objects.create(
        category=category,
        serial_no=serial,
        part_no=row.get('part_no'),
        brand=brand_name,
        model=row.get('model'),
        name=product_name or f"CONTROLLER - {category_name} - {row.get('model', '')} - {brand_name}".strip(' -'),
    )

    Spare.objects.create(
        product=product,
        controller=controller,
        brand=brand,
        model=row.get('model'),
        part_no=row.get('part_no'),
        alt_part_no=row.get('alt_part_no'),
        alt_serial_no=row.get('alt_serial_no'),
        specs=row.get('specs'),
        qty=row.get('qty') or 1,
        barcode=barcode or None,
        location=row.get('location') or controller.location,
        reference_location=row.get('reference_location') or controller.reference_location,
        remark=row.get('remark'),
    )

    last_controller_txn = InventoryTransaction.objects.filter(
        product=controller.product
    ).order_by('-created_at').first()

    from apps.core.services import _create_stock_in_transaction
    _create_stock_in_transaction(
        product=product,
        store_location=last_controller_txn.store_location if last_controller_txn else DEFAULT_STORE,
        stock_status=last_controller_txn.stock_status if last_controller_txn else DEFAULT_STATUS,
        user=user,
        remarks='Controller component import stock-in',
    )

    return product


def import_server_row(row, user=None):
    service_tag = (row.get('service_tag') or '').strip().upper()
    if not service_tag:
        raise ValueError('System Service Tag No is required')

    server_data = {
        'testing_date': row.get('testing_date'),
        'tested_by': row.get('tested_by'),
        'machine_type': row.get('machine_type'),
        'machine_no': row.get('machine_no'),
        'service_tag': service_tag,
        'model': row.get('server_model'),
        'brand': '',
        'part_no': '',
        'alt_part_no': '',
        'alt_serial_no': '',
        'specs': '',
        'qty': 1,
        'barcode': '',
        'status': 'WORKING',
        'location': row.get('location'),
        'reference_location': row.get('reference_location'),
        'parent_child_location': row.get('parent_child_location'),
        'remark': row.get('remark'),
        'store_location': DEFAULT_STORE,
        'stock_status': DEFAULT_STATUS,
    }
    component = {
        'spare_type': row.get('spare_type'),
        'brand': '',
        'model': row.get('server_model'),
        'part_no': row.get('part_no'),
        'alt_part_no': row.get('alt_part_no'),
        'serial_no': row.get('serial_no'),
        'alt_serial_no': row.get('alt_serial_no'),
        'specs': row.get('specs'),
        'barcode': row.get('barcode'),
        'qty': row.get('qty') or 1,
        'working_status': row.get('working_status') or 'WORKING',
        'location': row.get('location'),
        'reference_location': row.get('reference_location'),
        'parent_child_location': row.get('parent_child_location'),
        'remark': row.get('remark'),
    }

    server = Server.objects.filter(service_tag=service_tag).first()
    if not server:
        server = create_server_with_components(server_data, [component])
        mark_latest_stock_user(server.product, user)
    else:
        product = add_server_component(server, component, DEFAULT_STORE, DEFAULT_STATUS).product
        mark_latest_stock_user(product, user)
    return server.product
