from django.http import JsonResponse
from datetime import date, datetime

from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from apps.inventory.models import (
    AuditFinding,
    GeneralAuditFinding,
    InventoryFreezeRecord,
    InventoryTransaction,
    InventoryTransfer,
    RentalRecord,
    SalesReturn,
    TransferRequest,
    TransferRequestItem,
)
from apps.core.models import Product
from apps.core.activity import log_activity, log_timeline, format_dated_remark, append_dated_remark
from apps.core.models import ActivityLog, Notification
from apps.core.permissions import require_any_permission, require_permission
from django.utils.timezone import now
from django.db.models import Count, OuterRef, Subquery, Q
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from apps.core.permissions import has_permission
from openpyxl import load_workbook


def _product_for_barcode(barcode):
    """Resolve a physical item barcode across all category tables."""
    barcode = (barcode or '').strip()
    if not barcode:
        return None
    from apps.categories.models import Spare, Card, CPU, Controller, Memory, SFP, RailKit, HardDisk, NetworkingSpare
    from apps.servers.models import Server, ServerComponent
    for model in (Spare, Card, CPU, Controller, Memory, SFP, RailKit, HardDisk, NetworkingSpare, Server, ServerComponent):
        obj = model.objects.filter(barcode__iexact=barcode).select_related('product').first()
        if obj:
            return obj.product
    return None


def _latest_stock_out(product):
    return product.transactions.filter(transaction_type='OUT').order_by('-created_at', '-id').first()


def _append_remark_to_asset(product, remarks, when):
    formatted_remarks = format_dated_remark(remarks, when)
    if not formatted_remarks:
        return ''

    related_attrs = (
        'spare', 'card', 'cpu', 'controller', 'memory',
        'sfp', 'railkit', 'harddisk', 'networking_spare', 'server',
    )
    for attr in related_attrs:
        asset = getattr(product, attr, None)
        if asset is None or not hasattr(asset, 'remark'):
            continue
        current = getattr(asset, 'remark', '') or ''
        asset.remark = append_dated_remark(current, remarks, when)
        asset.save(update_fields=['remark'])
        break

    return formatted_remarks


def _create_return(product, stock_out, reason, disposition, remarks, user, returned_on, return_warehouse='', return_location=''):
    valid_locations = dict(InventoryTransaction.STORE_LOCATION)
    preferred_warehouse = (return_warehouse or '').strip().upper()
    final_warehouse = preferred_warehouse if preferred_warehouse in valid_locations else (stock_out.store_location or 'WH1')
    asset_location = (return_location or '').strip()
    formatted_remarks = _append_remark_to_asset(product, remarks, returned_on)
    SalesReturn.objects.create(
        product=product, stock_out_transaction=stock_out, reason=reason,
        disposition=disposition, remarks=formatted_remarks, returned_by=user, returned_on=returned_on,
    )
    InventoryTransaction.objects.create(
        product=product, transaction_type='IN', store_location=final_warehouse,
        stock_status=disposition, stock_in_date=returned_on, performed_by=user,
    )
    if asset_location:
        _set_product_location(product, asset_location)
    log_timeline(product=product, event_type='RETURN', user=user,
                 warehouse=final_warehouse, location=asset_location or _latest_location(product), remarks=formatted_remarks,
                 details={'reason': reason, 'disposition': disposition, 'stock_out_id': stock_out.id})


def _return_server_components(product, reason, disposition, remarks, user, returned_on, stock_status='SALE', return_warehouse='', return_location=''):
    server = getattr(product, 'server', None)
    if not server:
        return 0
    returned = 0
    for component in server.components.select_related('product'):
        component_out = _latest_stock_out(component.product)
        latest = component.product.transactions.order_by('-created_at', '-id').first()
        if component_out and latest and latest.id == component_out.id and component_out.stock_status == stock_status:
            _create_return(
                component.product, component_out, reason, disposition, remarks, user, returned_on,
                return_warehouse=return_warehouse, return_location=return_location,
            )
            returned += 1
    return returned


def _create_scrap_transaction(product, remarks, user, scrapped_on=None, scrap_location=''):
    latest = product.transactions.order_by('-created_at', '-id').first()
    valid_locations = dict(InventoryTransaction.STORE_LOCATION)
    preferred_location = (scrap_location or '').strip().upper()
    fallback_location = latest.store_location if latest else _latest_store_location(product)
    store_location = preferred_location if preferred_location in valid_locations else fallback_location
    formatted_remarks = _append_remark_to_asset(product, remarks, scrapped_on or date.today())
    txn = InventoryTransaction.objects.create(
        product=product,
        transaction_type='OUT',
        store_location=store_location or 'WH1',
        stock_status='SCRAP',
        stock_out_date=scrapped_on or date.today(),
        client_name=getattr(latest, 'client_name', '') or '',
        invoice_no=getattr(latest, 'invoice_no', '') or '',
        olf_dc_number=getattr(latest, 'olf_dc_number', '') or '',
        performed_by=user if getattr(user, 'is_authenticated', False) else None,
    )
    if formatted_remarks:
        log_timeline(
            product=product,
            event_type='SCRAP',
            user=user,
            warehouse=store_location or 'WH1',
            location=_latest_location(product),
            remarks=formatted_remarks,
            details={'auto_scrap': True},
        )
    return txn


def _scrap_server_components(product, remarks, user, scrapped_on=None):
    from apps.servers.models import Server

    server = Server.objects.filter(product=product).first()
    if not server:
        return 0

    count = 0
    for component in server.components.select_related('product').all():
        latest = component.product.transactions.order_by('-created_at', '-id').first()
        if latest and latest.stock_status in ('FAULTY', 'DAMAGED'):
            _create_scrap_transaction(component.product, remarks, user, scrapped_on=scrapped_on)
            count += 1
    return count

# Create your views here.


def _latest_store_location(product):
    last_txn = InventoryTransaction.objects.filter(
        product=product
    ).order_by('-created_at').first()

    return last_txn.store_location if last_txn else 'WH1'


_RELATED_ATTRS = (
    'spare', 'card', 'cpu', 'controller', 'memory',
    'sfp', 'railkit', 'harddisk', 'networking_spare', 'server',
)

OPERATIONAL_STOCK_OUT_STATUSES = (
    'TESTING', 'RENT', 'REPLACEMENT', 'ADV_REPLACEMENT',
    'ON_APPROVAL', 'EMPTY', 'REFILL', 'SCRAP',
)


def _related_inventory_objects(product):
    objects = []
    for attr in _RELATED_ATTRS:
        try:
            obj = getattr(product, attr, None)
        except Exception:
            obj = None
        if obj:
            objects.append(obj)
    return objects


def _latest_location(product):
    for obj in _related_inventory_objects(product):
        if getattr(obj, 'location', ''):
            return obj.location
    return ''


def _barcode_for_product(product):
    for obj in _related_inventory_objects(product):
        barcode = getattr(obj, 'barcode', '')
        if barcode:
            return barcode
    return product.serial_no or str(product.id)


def _set_product_location(product, location):
    """Update the location field on every related inventory record for *product*.

    Used by transfer so the asset's physical location follows the warehouse move.
    Returns the number of records updated.
    """
    updated = 0
    for obj in _related_inventory_objects(product):
        if hasattr(obj, 'location'):
            obj.location = location
            obj.save(update_fields=['location'])
            updated += 1
    return updated


def _is_frozen(product):
    latest = product.freeze_records.order_by('-frozen_at', '-id').first()
    return bool(latest and latest.status == 'FROZEN')


def _has_pending_transfer(product):
    return product.transfer_request_items.filter(received_at__isnull=True).exists()


def _create_stock_out_transaction(product, store_location, stock_status,
                                  stock_out_date, client_name, invoice_no,
                                  user=None, olf_dc_number=''):
    return InventoryTransaction.objects.create(
        product          = product,
        transaction_type = 'OUT',
        store_location   = store_location,
        stock_status     = stock_status,
        stock_out_date   = stock_out_date,
        client_name      = client_name,
        invoice_no       = invoice_no,
        olf_dc_number    = olf_dc_number,
        performed_by     = user if getattr(user, 'is_authenticated', False) else None,
    )


def create_rental_record(product, store_location, stock_out_date, client_name,
                         invoice_no, olf_dc_number, expected_return_date=None,
                         user=None, remarks=''):
    """Open a rental when an item is stocked out as RENT. Shared by the UI
    stock-out flow and the Excel stock-out import."""
    formatted_remarks = format_dated_remark(remarks, stock_out_date)
    return RentalRecord.objects.create(
        product=product,
        client_name=client_name or '',
        invoice_no=invoice_no or '',
        olf_dc_number=olf_dc_number or '',
        store_location=store_location or '',
        rent_out_date=stock_out_date,
        expected_return_date=expected_return_date,
        status='ON_RENT',
        rented_out_by=user if getattr(user, 'is_authenticated', False) else None,
        remarks=formatted_remarks or '',
    )


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), '%Y-%m-%d').date()
    except ValueError:
        return None


def _stock_out_server_components(product, store_location, stock_status,
                                 stock_out_date, client_name, invoice_no,
                                 user=None, olf_dc_number=''):
    from apps.servers.models import Server

    server = Server.objects.filter(product=product).first()
    if not server:
        return 0

    component_products = (
        server.components
        .select_related('product')
        .all()
    )

    count = 0
    for component in component_products:
        _create_stock_out_transaction(
            product=component.product,
            store_location=store_location,
            stock_status=stock_status,
            stock_out_date=stock_out_date,
            client_name=client_name,
            invoice_no=invoice_no,
            user=user,
            olf_dc_number=olf_dc_number,
        )
        count += 1

    return count
 
@require_permission('stock_out')
def stock_out(request):
    if request.method == 'POST':
 
        product_id   = request.POST.get('product_id')
        client_name  = request.POST.get('client_name', '')
        invoice_no   = request.POST.get('invoice_no', '')
        olf_dc_number = request.POST.get('olf_dc_number', '')
        stock_status = request.POST.get('stock_status', 'SALE')
        date_value   = request.POST.get('stock_out_date', '')
        expected_return = _parse_date(request.POST.get('expected_return_date', ''))
        # Frozen stock can only be disposed when explicitly stocked out from
        # the Frozen Stock page (Freeze module → "Mark Sold / Stock Out").
        allow_frozen = request.POST.get('allow_frozen') in ('1', 'true', 'True', 'on')

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Product not found'})
        was_frozen = _is_frozen(product)
        if was_frozen and not allow_frozen:
            return JsonResponse({'success': False, 'error': 'Frozen stock cannot be stocked out'})
        if _has_pending_transfer(product):
            return JsonResponse({'success': False, 'error': 'Product is pending receipt in a transfer request'})

        store_location = _latest_store_location(product)

        try:
            stock_out_date = (
                datetime.strptime(date_value, '%Y-%m-%d').date()
                if date_value else date.today()
            )
        except ValueError:
            stock_out_date = date.today()
 
        with transaction.atomic():
            if was_frozen and allow_frozen:
                # Disposing frozen stock releases the freeze so records stay consistent.
                latest_freeze = product.freeze_records.order_by('-frozen_at', '-id').first()
                InventoryFreezeRecord.objects.create(
                    product=product,
                    status='UNFROZEN',
                    reason=f'Released on stock out ({stock_status})',
                    frozen_by=latest_freeze.frozen_by if latest_freeze else None,
                    unfrozen_by=request.user,
                    unfrozen_at=now(),
                )
                log_timeline(
                    product=product,
                    event_type='UNFREEZE',
                    user=request.user,
                    warehouse=store_location,
                    remarks=f'Auto-unfrozen for stock out ({stock_status})',
                )

            _create_stock_out_transaction(
                product=product,
                store_location=store_location,
                stock_status=stock_status,
                stock_out_date=stock_out_date,
                client_name=client_name,
                invoice_no=invoice_no,
                user=request.user,
                olf_dc_number=olf_dc_number,
            )

            components_stocked_out = _stock_out_server_components(
                product=product,
                store_location=store_location,
                stock_status=stock_status,
                stock_out_date=stock_out_date,
                client_name=client_name,
                invoice_no=invoice_no,
                user=request.user,
                olf_dc_number=olf_dc_number,
            )

            if stock_status == 'RENT':
                create_rental_record(
                    product=product,
                    store_location=store_location,
                    stock_out_date=stock_out_date,
                    client_name=client_name,
                    invoice_no=invoice_no,
                    olf_dc_number=olf_dc_number,
                    expected_return_date=expected_return,
                    user=request.user,
                    remarks='Rented out',
                )

        log_activity(
            action='STOCK_OUT',
            module='INVENTORY',
            entity=product.name,
            entity_id=product.id,
            user=request.user,
            barcode='',
            warehouse=store_location,
            location=_latest_location(product),
            new_values={
                'stock_status': stock_status,
                'stock_out_date': str(stock_out_date),
                'client_name': client_name,
                'invoice_no': invoice_no,
                'olf_dc_number': olf_dc_number,
                'components_stocked_out': components_stocked_out,
            },
            remarks=f"Stock out for {product.serial_no or product.id}",
        )
        log_timeline(
            product=product,
            event_type='OUT',
            user=request.user,
            warehouse=store_location,
            remarks=client_name,
            details={
                'invoice_no': invoice_no,
                'olf_dc_number': olf_dc_number,
                'stock_status': stock_status,
            },
        )
 
        return JsonResponse({
            'success': True,
            'status': stock_status,
            'components_stocked_out': components_stocked_out,
        })
 
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@require_any_permission('sales_return', 'stock_return')
def sales_return(request):
    """Receive a stocked-out product. Normal returns become LIVE; faulty
    returns remain in a separate faulty register through their latest status."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    product = get_object_or_404(Product, id=request.POST.get('product_id'))
    reason = request.POST.get('reason', '').strip().upper()
    disposition = request.POST.get('disposition', '').strip().upper()
    return_warehouse = request.POST.get('location', '').strip().upper()
    return_location = request.POST.get('physical_location', '').strip()
    remarks = request.POST.get('remarks', '').strip()
    returned_on = _parse_date(request.POST.get('returned_on', '')) or date.today()
    if reason not in dict(SalesReturn.REASON_CHOICES):
        return JsonResponse({'success': False, 'error': 'Select a valid return reason'})
    if reason == 'NORMAL':
        disposition = 'LIVE'
    elif not disposition:
        disposition = 'FAULTY'
    elif disposition not in dict(SalesReturn.DISPOSITION_CHOICES):
        return JsonResponse({'success': False, 'error': 'Choose where the returned item should be stored'})
    latest = product.transactions.order_by('-created_at', '-id').first()
    stock_out = _latest_stock_out(product)
    if not stock_out or not latest or latest.id != stock_out.id:
        return JsonResponse({'success': False, 'error': 'Only an item currently stocked out can be returned'})
    with transaction.atomic():
        _create_return(
            product, stock_out, reason, disposition, remarks, request.user, returned_on,
            return_warehouse=return_warehouse,
            return_location=return_location,
        )
        components_returned = _return_server_components(
            product, reason, disposition, remarks, request.user, returned_on, stock_out.stock_status,
            return_warehouse=return_warehouse, return_location=return_location,
        )
    log_activity(action='SALES_RETURN', module='INVENTORY', entity=product.name, entity_id=product.id,
                 user=request.user, warehouse=stock_out.store_location, location=_latest_location(product),
                 new_values={'reason': reason, 'disposition': disposition, 'components_returned': components_returned},
                 remarks=format_dated_remark(remarks, returned_on) or 'Sales return processed')
    return JsonResponse({'success': True, 'disposition': disposition, 'components_returned': components_returned})


@require_permission('stock_out')
def scrap_faulty_stock(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    product = get_object_or_404(Product, id=request.POST.get('product_id'))
    remarks = request.POST.get('remarks', '').strip()
    scrap_location = request.POST.get('location', '').strip().upper()
    scrapped_on = _parse_date(request.POST.get('scrapped_on', '')) or date.today()
    latest = product.transactions.order_by('-created_at', '-id').first()
    if not latest or latest.stock_status not in ('FAULTY', 'DAMAGED'):
        return JsonResponse({'success': False, 'error': 'Only faulty or damaged stock can be scrapped'})
    with transaction.atomic():
        scrap_txn = _create_scrap_transaction(
            product, remarks, request.user, scrapped_on=scrapped_on, scrap_location=scrap_location
        )
        components_scrapped = _scrap_server_components(product, remarks, request.user, scrapped_on=scrapped_on)
    formatted_remarks = format_dated_remark(remarks, scrapped_on)
    log_activity(
        action='SCRAP',
        module='INVENTORY',
        entity=product.name,
        entity_id=product.id,
        user=request.user,
        barcode='',
        warehouse=scrap_txn.store_location,
        location=_latest_location(product),
        new_values={
            'stock_status': 'SCRAP',
            'stock_out_date': str(scrapped_on),
            'components_scrapped': components_scrapped,
        },
        remarks=formatted_remarks or 'Faulty stock scrapped',
    )
    log_timeline(
        product=product,
        event_type='SCRAP',
        user=request.user,
        warehouse=scrap_txn.store_location,
        location=_latest_location(product),
        remarks=formatted_remarks or 'Faulty stock scrapped',
        details={'components_scrapped': components_scrapped},
    )
    return JsonResponse({'success': True, 'status': 'SCRAP', 'components_scrapped': components_scrapped})


@require_permission('sales_return')
def sales_return_history(request):
    records = SalesReturn.objects.select_related('product', 'returned_by').all()
    q = request.GET.get('q', '').strip()
    if q:
        records = records.filter(Q(product__name__icontains=q) | Q(product__serial_no__icontains=q) | Q(remarks__icontains=q))
    return render(request, 'inventory/sales_return_list.html', {'records': records[:2000], 'q': q})


@require_any_permission('sold_view', 'reports')
def stock_out_status_list(request, status):
    status = (status or '').strip().upper()
    status_labels = dict(InventoryTransaction.STOCK_STATUS)
    if status not in OPERATIONAL_STOCK_OUT_STATUSES:
        return JsonResponse({'success': False, 'error': 'Unknown stock-out status'}, status=404)

    latest_txn = InventoryTransaction.objects.filter(product_id=OuterRef('id')).order_by('-created_at', '-id')
    products = Product.objects.select_related('category').annotate(
        latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
        latest_status=Subquery(latest_txn.values('stock_status')[:1]),
        latest_store=Subquery(latest_txn.values('store_location')[:1]),
        latest_client=Subquery(latest_txn.values('client_name')[:1]),
        latest_invoice=Subquery(latest_txn.values('invoice_no')[:1]),
        latest_olf_dc=Subquery(latest_txn.values('olf_dc_number')[:1]),
        latest_out_date=Subquery(latest_txn.values('stock_out_date')[:1]),
        latest_created_at=Subquery(latest_txn.values('created_at')[:1]),
    ).filter(latest_type='OUT', latest_status=status)

    q = request.GET.get('q', '').strip()
    if q:
        products = products.filter(
            Q(name__icontains=q) |
            Q(serial_no__icontains=q) |
            Q(part_no__icontains=q) |
            Q(latest_client__icontains=q) |
            Q(latest_invoice__icontains=q) |
            Q(latest_olf_dc__icontains=q)
        )

    rows = []
    for product in products[:2000]:
        detail = _inventory_detail(product)
        rows.append({
            'product': product,
            'category': product.category.name if product.category else '',
            'part_no': detail.get('part_no') or product.part_no or '',
            'barcode': detail.get('barcode') or '',
        })

    return render(request, 'inventory/stock_out_status_list.html', {
        'rows': rows,
        'q': q,
        'status': status,
        'status_label': status_labels.get(status, status),
        'statuses': [(value, status_labels.get(value, value)) for value in OPERATIONAL_STOCK_OUT_STATUSES],
        'result_count': len(rows),
        'can_stock_return': has_permission(request.user, 'stock_return'),
        'can_stock_out': has_permission(request.user, 'stock_out'),
    })



@require_permission('audit')
def audit_spare(request):
    if request.method == 'POST':
 
        product_id   = request.POST.get('product_id')
        audit_remark = request.POST.get('audit_remark', '')
        audited_on   = request.POST.get('audited_on', '')
        audit_result = request.POST.get('audit_result', 'FOUND')
 
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Product not found'})
        if _is_frozen(product):
            return JsonResponse({'success': False, 'error': 'Frozen stock cannot be audited'})
 
        last_txn = InventoryTransaction.objects.filter(
            product=product
        ).order_by('-created_at').first()
 
        InventoryTransaction.objects.create(
            product          = product,
            transaction_type = 'AUDIT',
            store_location   = last_txn.store_location if last_txn else 'WH1',
            stock_status     = last_txn.stock_status   if last_txn else 'LIVE',
            audited_on       = audited_on or now().date(),
            audited_by       = request.user if request.user.is_authenticated else None,
            performed_by     = request.user if request.user.is_authenticated else None,
            audit_remark     = audit_remark,
            audit_result     = audit_result,
        )

        log_activity(
            action='AUDIT',
            module='AUDIT',
            entity=product.name,
            entity_id=product.id,
            user=request.user,
            warehouse=last_txn.store_location if last_txn else 'WH1',
            new_values={
                'audited_on': str(audited_on or now().date()),
                'audit_result': audit_result,
            },
            remarks=audit_remark,
        )
        log_timeline(
            product=product,
            event_type='AUDIT',
            user=request.user,
            warehouse=last_txn.store_location if last_txn else 'WH1',
            remarks=audit_remark,
            details={'audit_result': audit_result},
        )
 
        return JsonResponse({'success': True})
 
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@require_permission('product_history')
def audit_history(request, product_id):

    audits = InventoryTransaction.objects.filter(
        product_id=product_id,
        transaction_type='AUDIT'
    ).order_by('-created_at')

    data = []

    for a in audits:
        data.append({
            "date": a.audited_on,
            "user": a.audited_by.username if a.audited_by else "",
            "location": a.store_location,
            "status": a.stock_status,
            "remark": a.audit_remark,
            "audit_result": a.audit_result or '',
        })

    return JsonResponse({"data": data})


@require_permission('audit_view')
def audit_report(request):
    base = InventoryTransaction.objects.filter(
        transaction_type='AUDIT'
    ).select_related('product', 'product__category', 'audited_by')
    total_count = base.count()

    audits = base.order_by('-created_at')

    q = request.GET.get('q', '').strip()
    user_id = request.GET.get('user', '').strip()
    result = request.GET.get('result', '').strip()
    category = request.GET.get('category', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if q:
        audits = audits.filter(
            Q(product__name__icontains=q) |
            Q(product__serial_no__icontains=q) |
            Q(product__part_no__icontains=q) |
            Q(audit_remark__icontains=q) |
            Q(store_location__icontains=q)
        )
    if user_id:
        audits = audits.filter(audited_by_id=user_id)
    if result:
        audits = audits.filter(audit_result=result)
    if category:
        audits = audits.filter(product__category__name=category)
    if date_from:
        audits = audits.filter(audited_on__gte=date_from)
    if date_to:
        audits = audits.filter(audited_on__lte=date_to)

    audits = list(audits[:2000])

    from apps.core.models import SpareCategory
    audit_user_ids = base.exclude(audited_by__isnull=True).values_list('audited_by', flat=True).distinct()
    from django.contrib.auth.models import User as AuthUser
    audit_users = AuthUser.objects.filter(id__in=list(audit_user_ids)).order_by('username')

    return render(request, 'spare/audit_report.html', {
        'audits': audits,
        'total_count': total_count,
        'result_count': len(audits),
        'audit_users': audit_users,
        'categories': SpareCategory.objects.order_by('name'),
        'result_choices': InventoryTransaction.AUDIT_RESULT,
        'filters': {
            'q': q, 'user': user_id, 'result': result, 'category': category,
            'date_from': date_from, 'date_to': date_to,
        },
    })
    
    

def check_product_membership(request):
    """
    Given a product_id, returns whether the product is part of
    a server or controller, with their labels.
    Used to show warning before stock-out / audit.
    """
    from apps.servers.models import ServerComponent
    from apps.categories.models import Spare
 
    product_id = request.GET.get('product_id')
    if not product_id:
        return JsonResponse({'in_server': False, 'in_controller': False})
 
    result = {'in_server': False, 'in_controller': False,
              'server_label': '', 'controller_label': ''}
 
    # check server
    sc = ServerComponent.objects.filter(
        product_id=product_id
    ).select_related('server').first()
 
    if sc:
        result['in_server']     = True
        result['server_label']  = (
            f"{sc.server.machine_no or ''} — "
            f"{sc.server.model or ''} "
            f"[{sc.server.service_tag}]"
        ).strip(' —')
 
    # check controller (via Spare.controller FK)
    try:
        spare = Spare.objects.select_related('controller__product').get(
            product_id=product_id
        )
        if spare.controller:
            result['in_controller']    = True
            result['controller_label'] = (
                f"{spare.controller.model or ''} "
                f"[{spare.controller.product.serial_no}]"
            )
    except Spare.DoesNotExist:
        pass
 
    return JsonResponse(result)


@require_permission('transfer_request')
def transfer_inventory(request):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=request.POST.get('product_id'))
        if _is_frozen(product):
            return JsonResponse({'success': False, 'error': 'Cannot transfer frozen stock'})

        destination_warehouse = request.POST.get('destination_warehouse', '').strip().upper()
        destination_location = request.POST.get('destination_location', '').strip()
        remarks = request.POST.get('remarks', '').strip()
        if not remarks:
            return JsonResponse({'success': False, 'error': 'Remarks are required'})
        valid_warehouses = dict(InventoryTransaction.STORE_LOCATION)
        if destination_warehouse not in valid_warehouses:
            return JsonResponse({'success': False, 'error': 'Destination warehouse must exist'})

        source_warehouse = _latest_store_location(product)
        source_location = _latest_location(product)
        latest = product.transactions.order_by('-created_at', '-id').first()
        if _has_pending_transfer(product) or not latest or latest.transaction_type != 'IN':
            return JsonResponse({'success': False, 'error': 'Product is not available for a new transfer request'})
        with transaction.atomic():
            transfer_request = TransferRequest.objects.create(
                source_warehouse=source_warehouse,
                destination_warehouse=destination_warehouse,
                remarks=format_dated_remark(remarks) or remarks,
                requested_by=request.user,
            )
            TransferRequestItem.objects.create(
                request=transfer_request,
                product=product,
                barcode=_barcode_for_product(product),
                source_location=source_location,
                destination_location='',
            )
        log_activity(
            action='TRANSFER_REQUEST', module='INVENTORY', entity=product.name,
            entity_id=product.id, user=request.user, warehouse=source_warehouse,
            location=source_location,
            new_values={'transfer_request_id': transfer_request.id, 'destination_warehouse': destination_warehouse},
            remarks=format_dated_remark(remarks) or remarks,
        )
        log_timeline(
            product=product, event_type='TRANSFER', user=request.user,
            warehouse=source_warehouse, location=source_location,
            remarks=format_dated_remark(f'Transfer request TR-{transfer_request.id}: {remarks}') or remarks,
            details={'destination_warehouse': destination_warehouse, 'pending_receipt': True},
        )
        return JsonResponse({'success': True, 'transfer_request_id': transfer_request.id, 'message': 'Transfer request created. It must be received by Stock In or Admin.'})

        with transaction.atomic():
            _set_product_location(product, destination_location)
            transfer = InventoryTransfer.objects.create(
                product=product,
                source_warehouse=source_warehouse,
                destination_warehouse=destination_warehouse,
                source_location=source_location,
                destination_location=destination_location,
                remarks=format_dated_remark(remarks) or remarks,
                transferred_by=request.user,
            )
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='AUDIT',
                store_location=destination_warehouse,
                stock_status='LIVE',
                audited_on=now().date(),
                audited_by=request.user,
                performed_by=request.user,
                audit_remark=f"Transferred from {source_warehouse} to {destination_warehouse}. {remarks}",
                audit_result='FOUND',
            )
            # 🔥 Move the physical location too (the actual bug fix):
            # the new warehouse is recorded on the transaction, but the
            # asset's location lives on the related category record.
            _set_product_location(product, destination_location)
        log_activity(
            action='TRANSFER',
            module='INVENTORY',
            entity=product.name,
            entity_id=product.id,
            user=request.user,
            warehouse=destination_warehouse,
            location=destination_location,
            old_values={'warehouse': source_warehouse, 'location': source_location},
            new_values={'warehouse': destination_warehouse, 'location': destination_location, 'transfer_id': transfer.id},
            remarks=format_dated_remark(remarks) or remarks,
        )
        log_timeline(
            product=product,
            event_type='TRANSFER',
            user=request.user,
            warehouse=destination_warehouse,
            location=destination_location,
            remarks=format_dated_remark(remarks) or remarks,
            details={'source_warehouse': source_warehouse, 'source_location': source_location},
        )
        Notification.objects.create(
            user=request.user,
            notification_type='TRANSFER',
            title='Inventory transferred',
            message=f'{product.serial_no or product.id} moved to {destination_warehouse}',
            metadata={'product_id': product.id, 'transfer_id': transfer.id},
        )
        return JsonResponse({'success': True, 'transfer_id': transfer.id})
    return JsonResponse({'success': False, 'error': 'Invalid method'})


def _transfer_sheet_rows(upload):
    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or '').strip().lower() for value in next(iterator)]
    except Exception as exc:
        raise ValueError(f'Unable to read Excel file: {exc}')
    barcode_index = next((i for i, name in enumerate(headers) if name in ('barcode', 'barcode no', 'barcode no.', 'barcode number')), None)
    location_index = next((i for i, name in enumerate(headers) if name in ('location', 'updated location', 'destination location')), None)
    if barcode_index is None:
        raise ValueError('Excel needs a Barcode or Barcode No column')
    for row_no, row in enumerate(iterator, start=2):
        barcode = str(row[barcode_index] or '').strip() if barcode_index < len(row) else ''
        location = str(row[location_index] or '').strip() if location_index is not None and location_index < len(row) else ''
        if barcode:
            yield row_no, barcode, location


@login_required
def transfer_request_page(request):
    if not (has_permission(request.user, 'transfer_request') or has_permission(request.user, 'transfer_receive')):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    requests = list(TransferRequest.objects.select_related('requested_by').prefetch_related('items__product').annotate(
        received_count=Count('items', filter=Q(items__received_at__isnull=False))
    )[:500])
    for transfer_request in requests:
        for item in transfer_request.items.all():
            detail = _inventory_detail(item.product)
            item.display_part_no = detail.get('part_no') or item.product.part_no or ''
            item.display_barcode = item.barcode or detail.get('barcode') or item.product.serial_no or ''
    pending_requests = [req for req in requests if req.status == 'PENDING']
    completed_requests = [req for req in requests if req.status in ('PARTIAL', 'COMPLETED')]
    return render(request, 'inventory/transfer_requests.html', {
        'pending_requests': pending_requests,
        'completed_requests': completed_requests,
        'receivable_requests': [req for req in requests if req.status in ('PENDING', 'PARTIAL')],
        'warehouses': InventoryTransaction.STORE_LOCATION,
    })


@login_required
def create_transfer_request(request):
    if not has_permission(request.user, 'transfer_request'):
        return JsonResponse({'success': False, 'error': 'Only Stock Out users or Admin can request a transfer'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    source = request.POST.get('source_warehouse', '').strip().upper()
    destination = request.POST.get('destination_warehouse', '').strip().upper()
    upload = request.FILES.get('file')
    if source not in dict(InventoryTransaction.STORE_LOCATION) or destination not in dict(InventoryTransaction.STORE_LOCATION):
        return JsonResponse({'success': False, 'error': 'Select valid source and destination stores'})
    if source == destination:
        return JsonResponse({'success': False, 'error': 'Source and destination stores must differ'})
    if not upload:
        return JsonResponse({'success': False, 'error': 'Upload an Excel file containing barcodes'})
    errors, seen = [], set()
    with transaction.atomic():
        formatted_remarks = format_dated_remark(request.POST.get('remarks', '').strip())
        transfer_request = TransferRequest.objects.create(source_warehouse=source, destination_warehouse=destination,
            remarks=formatted_remarks, requested_by=request.user)
        try:
            rows = _transfer_sheet_rows(upload)
            for row_no, barcode, _ in rows:
                barcode_key = barcode.upper()
                if barcode_key in seen:
                    errors.append(f'Row {row_no}: duplicate barcode')
                    continue
                seen.add(barcode_key)
                product = _product_for_barcode(barcode)
                latest = product.transactions.order_by('-created_at', '-id').first() if product else None
                if not product:
                    errors.append(f'Row {row_no}: barcode not found ({barcode})')
                elif _is_frozen(product):
                    errors.append(f'Row {row_no}: product is frozen ({barcode})')
                elif _has_pending_transfer(product):
                    errors.append(f'Row {row_no}: product is already pending transfer ({barcode})')
                elif not latest or latest.transaction_type != 'IN':
                    errors.append(f'Row {row_no}: product is not currently in stock ({barcode})')
                elif latest.store_location != source:
                    errors.append(f'Row {row_no}: product is not in source store {source} ({barcode})')
                else:
                    TransferRequestItem.objects.create(request=transfer_request, product=product, barcode=barcode, source_location=_latest_location(product))
        except ValueError as exc:
            transfer_request.delete()
            return JsonResponse({'success': False, 'error': str(exc)})
        if not transfer_request.items.exists():
            transfer_request.delete()
            return JsonResponse({'success': False, 'error': 'No eligible barcodes found', 'errors': errors})
    return JsonResponse({'success': True, 'request_id': transfer_request.id, 'accepted': transfer_request.items.count(), 'errors': errors})


@login_required
def receive_transfer_request(request, request_id):
    if not has_permission(request.user, 'transfer_receive'):
        return JsonResponse({'success': False, 'error': 'Only Stock In users or Admin can receive a transfer'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'success': False, 'error': 'Upload an Excel file with Barcode and updated Location columns'})
    transfer_request = get_object_or_404(TransferRequest, id=request_id)
    pending = {item.barcode.upper(): item for item in transfer_request.items.select_related('product').filter(received_at__isnull=True)}
    errors, received = [], 0
    with transaction.atomic():
        try:
            rows = _transfer_sheet_rows(upload)
            for row_no, barcode, destination_location in rows:
                item = pending.get(barcode.upper())
                if not item:
                    errors.append(f'Row {row_no}: barcode is not pending on this transfer request ({barcode})')
                    continue
                if not destination_location:
                    errors.append(f'Row {row_no}: updated location is required ({barcode})')
                    continue
                product = item.product
                _set_product_location(product, destination_location)
                InventoryTransfer.objects.create(product=product, source_warehouse=transfer_request.source_warehouse,
                    destination_warehouse=transfer_request.destination_warehouse, source_location=item.source_location,
                    destination_location=destination_location,
                    remarks=format_dated_remark(f'Transfer request TR-{transfer_request.id}', date.today()),
                    transferred_by=request.user)
                InventoryTransaction.objects.create(product=product, transaction_type='IN', store_location=transfer_request.destination_warehouse,
                    stock_status='LIVE', stock_in_date=date.today(), performed_by=request.user)
                item.destination_location = destination_location
                item.received_by = request.user
                item.received_at = now()
                item.save(update_fields=['destination_location', 'received_by', 'received_at'])
                log_timeline(product=product, event_type='TRANSFER', user=request.user, warehouse=transfer_request.destination_warehouse,
                             location=destination_location, remarks=f'Received against TR-{transfer_request.id}')
                received += 1
        except ValueError as exc:
            return JsonResponse({'success': False, 'error': str(exc)})
        pending_count = transfer_request.items.filter(received_at__isnull=True).count()
        transfer_request.status = 'COMPLETED' if not pending_count else ('PARTIAL' if received else 'PENDING')
        transfer_request.completed_at = now() if not pending_count else None
        transfer_request.save(update_fields=['status', 'completed_at'])
    return JsonResponse({'success': True, 'received': received, 'pending': pending_count, 'errors': errors})

@login_required
def receive_transfer_item(request, request_id, item_id):
    """Approve one requested item without an Excel upload."""
    if not has_permission(request.user, 'transfer_receive'):
        return JsonResponse({'success': False, 'error': 'Only Stock In users or Admin can receive a transfer'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    destination_location = request.POST.get('destination_location', '').strip()
    if not destination_location:
        return JsonResponse({'success': False, 'error': 'Updated location is required'})
    item = get_object_or_404(TransferRequestItem.objects.select_related('request', 'product'), id=item_id, request_id=request_id)
    if item.received_at:
        return JsonResponse({'success': False, 'error': 'This item has already been received'})
    transfer_request = item.request
    with transaction.atomic():
        _set_product_location(item.product, destination_location)
        InventoryTransfer.objects.create(
            product=item.product, source_warehouse=transfer_request.source_warehouse,
            destination_warehouse=transfer_request.destination_warehouse,
            source_location=item.source_location, destination_location=destination_location,
            remarks=format_dated_remark(f'Transfer request TR-{transfer_request.id}', date.today()), transferred_by=request.user,
        )
        InventoryTransaction.objects.create(
            product=item.product, transaction_type='IN', store_location=transfer_request.destination_warehouse,
            stock_status='LIVE', stock_in_date=date.today(), performed_by=request.user,
        )
        item.destination_location = destination_location
        item.received_by = request.user
        item.received_at = now()
        item.save(update_fields=['destination_location', 'received_by', 'received_at'])
        pending_count = transfer_request.items.filter(received_at__isnull=True).count()
        transfer_request.status = 'COMPLETED' if not pending_count else 'PARTIAL'
        transfer_request.completed_at = now() if not pending_count else None
        transfer_request.save(update_fields=['status', 'completed_at'])
    log_activity(action='TRANSFER_RECEIVED', module='INVENTORY', entity=item.product.name, entity_id=item.product.id,
                 user=request.user, warehouse=transfer_request.destination_warehouse, location=destination_location,
                 new_values={'transfer_request_id': transfer_request.id}, remarks='Single-item transfer receipt')
    log_timeline(product=item.product, event_type='TRANSFER', user=request.user,
                 warehouse=transfer_request.destination_warehouse, location=destination_location,
                 remarks=f'Received against TR-{transfer_request.id}')
    return JsonResponse({'success': True, 'pending': pending_count})


@require_permission('product_history')
def transfer_history(request):
    base = InventoryTransfer.objects.select_related('product', 'transferred_by')
    total_count = base.count()
    transfers = base.order_by('-created_at')

    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if q:
        transfers = transfers.filter(
            Q(product__name__icontains=q) |
            Q(product__serial_no__icontains=q) |
            Q(source_warehouse__icontains=q) |
            Q(destination_warehouse__icontains=q) |
            Q(destination_location__icontains=q) |
            Q(remarks__icontains=q)
        )
    if date_from:
        transfers = transfers.filter(created_at__date__gte=date_from)
    if date_to:
        transfers = transfers.filter(created_at__date__lte=date_to)

    transfers = list(transfers)
    return render(request, 'inventory/transfer_history.html', {
        'transfers': transfers,
        'total_count': total_count,
        'result_count': len(transfers),
        'filters': {'q': q, 'date_from': date_from, 'date_to': date_to},
    })


def _freeze_related_children(product, user, reason):
    frozen = 0
    server = getattr(product, 'server', None)
    if server:
        for component in server.components.select_related('product'):
            InventoryFreezeRecord.objects.create(
                product=component.product,
                status='FROZEN',
                reason=reason,
                frozen_by=user,
                parent_product=product,
            )
            log_timeline(product=component.product, event_type='FREEZE', user=user, remarks=reason)
            frozen += 1
    controller = getattr(product, 'controller', None)
    if controller:
        for component in controller.components.select_related('product'):
            InventoryFreezeRecord.objects.create(
                product=component.product,
                status='FROZEN',
                reason=reason,
                frozen_by=user,
                parent_product=product,
            )
            log_timeline(product=component.product, event_type='FREEZE', user=user, remarks=reason)
            frozen += 1
    return frozen


@require_permission('freeze')
def freeze_inventory(request):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=request.POST.get('product_id'))
        reason = request.POST.get('reason', '').strip()
        if not reason:
            return JsonResponse({'success': False, 'error': 'Remarks are required'})
        if _is_frozen(product):
            return JsonResponse({'success': False, 'error': 'Asset is already frozen'})
        record = InventoryFreezeRecord.objects.create(
            product=product,
            status='FROZEN',
            reason=reason,
            frozen_by=request.user,
        )
        affected = _freeze_related_children(product, request.user, reason)
        log_activity(
            action='FREEZE',
            module='INVENTORY',
            entity=product.name,
            entity_id=product.id,
            user=request.user,
            warehouse=_latest_store_location(product),
            location=_latest_location(product),
            new_values={'freeze_id': record.id, 'affected_components': affected},
            remarks=reason,
        )
        log_timeline(product=product, event_type='FREEZE', user=request.user, warehouse=_latest_store_location(product), location=_latest_location(product), remarks=reason)
        Notification.objects.create(
            user=request.user,
            notification_type='FREEZE',
            title='Inventory frozen',
            message=f'{product.serial_no or product.id} frozen',
            metadata={'product_id': product.id, 'freeze_id': record.id},
        )
        return JsonResponse({'success': True, 'affected_components': affected})
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@require_permission('freeze')
def unfreeze_inventory(request):
    if request.method == 'POST':
        wants_json = request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('accept', '')
        product = get_object_or_404(Product, id=request.POST.get('product_id'))
        reason = request.POST.get('reason', '').strip()
        if not reason:
            if wants_json:
                return JsonResponse({'success': False, 'error': 'Remarks are required'})
            return redirect('frozen_inventory_list')
        latest = product.freeze_records.order_by('-frozen_at', '-id').first()
        if not latest or latest.status != 'FROZEN':
            if wants_json:
                return JsonResponse({'success': False, 'error': 'Asset is not frozen'})
            return redirect('frozen_inventory_list')

        # Disposition + warehouse/location, mirroring the Sales Return flow.
        disposition = (request.POST.get('disposition') or 'LIVE').strip().upper()
        if disposition not in {'LIVE', 'FAULTY', 'DAMAGED', 'SCRAP'}:
            disposition = 'LIVE'
        return_warehouse = (request.POST.get('return_warehouse') or '').strip().upper()
        return_location = (request.POST.get('return_location') or '').strip()
        valid_locations = dict(InventoryTransaction.STORE_LOCATION)
        store_location = (
            return_warehouse if return_warehouse in valid_locations
            else _latest_store_location(product)
        )

        with transaction.atomic():
            InventoryFreezeRecord.objects.create(
                product=product,
                status='UNFROZEN',
                reason=reason,
                frozen_by=latest.frozen_by,
                unfrozen_by=request.user,
                unfrozen_at=now(),
            )
            # Optionally change disposition on unfreeze.
            if disposition == 'SCRAP':
                _create_scrap_transaction(product, reason, request.user, scrap_location=store_location)
            elif disposition in {'FAULTY', 'DAMAGED'}:
                InventoryTransaction.objects.create(
                    product=product, transaction_type='IN', store_location=store_location,
                    stock_status=disposition, stock_in_date=now().date(),
                    performed_by=request.user if request.user.is_authenticated else None,
                )
            if return_location:
                _set_product_location(product, return_location)

        log_activity(
            action='UNFREEZE',
            module='INVENTORY',
            entity=product.name,
            entity_id=product.id,
            user=request.user,
            warehouse=store_location,
            location=return_location or _latest_location(product),
            new_values={'disposition': disposition},
            remarks=reason,
        )
        log_timeline(product=product, event_type='UNFREEZE', user=request.user, warehouse=store_location, location=return_location or _latest_location(product), remarks=reason)
        if wants_json:
            return JsonResponse({'success': True})
        return redirect('frozen_inventory_list')
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@require_permission('freeze')
def frozen_inventory_list(request):
    latest = InventoryFreezeRecord.objects.filter(
        product=OuterRef('pk')
    ).order_by('-frozen_at', '-id')
    products = Product.objects.annotate(
        freeze_status=Subquery(latest.values('status')[:1]),
        freeze_reason=Subquery(latest.values('reason')[:1]),
        frozen_by_name=Subquery(latest.values('frozen_by__username')[:1]),
        frozen_at=Subquery(latest.values('frozen_at')[:1]),
    ).filter(freeze_status='FROZEN').order_by('-id')

    q = request.GET.get('q', '').strip()
    if q:
        products = products.filter(
            Q(name__icontains=q) |
            Q(serial_no__icontains=q) |
            Q(part_no__icontains=q) |
            Q(freeze_reason__icontains=q)
        )

    total_frozen = Product.objects.annotate(
        freeze_status=Subquery(latest.values('status')[:1]),
    ).filter(freeze_status='FROZEN').count()

    products = list(products)
    return render(request, 'inventory/frozen_list.html', {
        'products': products,
        'q': q,
        'result_count': len(products),
        'total_frozen': total_frozen,
    })


@require_permission('audit_findings')
def audit_finding_list(request):
    base = AuditFinding.objects.select_related('created_by')
    total_count = base.count()
    findings = base.order_by('-audit_date', '-id')

    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if q:
        findings = findings.filter(
            Q(remarks__icontains=q) | Q(person_involved__icontains=q)
        )
    if date_from:
        findings = findings.filter(audit_date__gte=date_from)
    if date_to:
        findings = findings.filter(audit_date__lte=date_to)

    findings = list(findings)
    return render(request, 'inventory/audit_findings.html', {
        'findings': findings,
        'total_count': total_count,
        'result_count': len(findings),
        'filters': {'q': q, 'date_from': date_from, 'date_to': date_to},
    })


@require_permission('audit')
def create_audit_finding(request):
    if request.method == 'POST':
        finding = AuditFinding.objects.create(
            audit_date=request.POST.get('audit_date') or now().date(),
            remarks=format_dated_remark(request.POST.get('remarks', '').strip()) or request.POST.get('remarks', '').strip(),
            attachment=request.FILES.get('attachment'),
            person_involved=request.POST.get('person_involved', '').strip(),
            created_by=request.user,
        )
        log_activity(
            action='AUDIT_FINDING_CREATE',
            module='AUDIT',
            entity='AuditFinding',
            entity_id=finding.id,
            user=request.user,
            remarks=finding.remarks,
        )
        Notification.objects.create(
            user=request.user,
            notification_type='AUDIT_FINDING',
            title='Audit finding created',
            message=f'Audit finding #{finding.id} created',
            metadata={'finding_id': finding.id},
        )
        return redirect('audit_finding_list')
    return render(request, 'inventory/audit_finding_form.html')


@require_permission('audit_findings')
def general_audit_finding_list(request):
    base = GeneralAuditFinding.objects.select_related('created_by', 'attended_by')
    total_count = base.count()
    findings = base.order_by('-audit_date', '-id')

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if q:
        findings = findings.filter(
            Q(title__icontains=q) | Q(remarks__icontains=q) | Q(person__icontains=q)
        )
    if status:
        findings = findings.filter(status=status)
    if date_from:
        findings = findings.filter(audit_date__gte=date_from)
    if date_to:
        findings = findings.filter(audit_date__lte=date_to)

    findings = list(findings)
    return render(request, 'inventory/general_audit_findings.html', {
        'findings': findings,
        'total_count': total_count,
        'result_count': len(findings),
        'status_choices': GeneralAuditFinding.STATUS_CHOICES,
        'filters': {'q': q, 'status': status, 'date_from': date_from, 'date_to': date_to},
    })


@require_permission('audit')
def create_general_audit_finding(request):
    if request.method == 'POST':
        finding = GeneralAuditFinding.objects.create(
            audit_date=request.POST.get('audit_date') or now().date(),
            title=request.POST.get('title', '').strip(),
            remarks=format_dated_remark(request.POST.get('remarks', '').strip()) or request.POST.get('remarks', '').strip(),
            attachment=request.FILES.get('attachment'),
            person=request.POST.get('person', '').strip(),
            created_by=request.user,
        )
        log_activity(
            action='GENERAL_AUDIT_FINDING_CREATE',
            module='AUDIT',
            entity='GeneralAuditFinding',
            entity_id=finding.id,
            user=request.user,
            remarks=finding.remarks,
        )
        return redirect('general_audit_finding_list')
    return render(request, 'inventory/general_audit_finding_form.html')


@require_permission('attend_audit_finding')
def attend_general_audit_finding(request, finding_id):
    finding = get_object_or_404(GeneralAuditFinding, id=finding_id)
    if request.method == 'POST':
        remarks = request.POST.get('remarks', '').strip()
        if not remarks:
            return redirect('general_audit_finding_list')
        finding.status = 'ATTENDED'
        finding.attended_by = request.user
        finding.attended_at = now()
        finding.attended_remarks = format_dated_remark(remarks) or remarks
        update_fields = ['status', 'attended_by', 'attended_at', 'attended_remarks']
        attachment = request.FILES.get('attended_attachment')
        if attachment:
            finding.attended_attachment = attachment
            update_fields.append('attended_attachment')
        finding.save(update_fields=update_fields)
        log_activity(
            action='GENERAL_AUDIT_FINDING_ATTENDED',
            module='AUDIT',
            entity='GeneralAuditFinding',
            entity_id=finding.id,
            user=request.user,
            remarks=format_dated_remark(remarks) or remarks,
        )
        return redirect('general_audit_finding_list')
    return redirect('general_audit_finding_list')


@require_permission('reports')
def activity_trail(request):
    logs = ActivityLog.objects.select_related('user').order_by('-timestamp')
    q = request.GET.get('q', '').strip()
    action = request.GET.get('action', '').strip()
    user_id = request.GET.get('user', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if q:
        logs = logs.filter(
            Q(barcode__icontains=q) |
            Q(entity__icontains=q) |
            Q(action__icontains=q) |
            Q(remarks__icontains=q) |
            Q(location__icontains=q) |
            Q(warehouse__icontains=q)
        )
    if action:
        logs = logs.filter(action=action)
    if user_id:
        logs = logs.filter(user_id=user_id)
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    actions = ActivityLog.objects.order_by('action').values_list('action', flat=True).distinct()
    users = ActivityLog.objects.exclude(user__isnull=True).select_related('user').order_by('user__username')
    unique_users = []
    seen = set()
    for log in users:
        if log.user_id in seen:
            continue
        seen.add(log.user_id)
        unique_users.append(log.user)

    total_count = ActivityLog.objects.count()
    result_count = logs.count()
    has_filters = bool(q or action or user_id or date_from or date_to)

    return render(request, 'inventory/activity_trail.html', {
        'logs': logs[:1000],
        'actions': actions,
        'activity_users': unique_users,
        'total_count': total_count,
        'result_count': result_count,
        'has_filters': has_filters,
    })


@require_permission('product_history')
def asset_movement_ledger(request, product_id):
    """Full product detail: current state + complete history — rent, sale,
    stock-in, audit, transfers, freeze/unfreeze, mapping, timeline, activity."""
    product = get_object_or_404(Product, id=product_id)

    txns = (InventoryTransaction.objects
            .filter(product=product)
            .select_related('performed_by', 'audited_by')
            .order_by('-created_at'))
    stock_in = [t for t in txns if t.transaction_type == 'IN']
    stock_out = [t for t in txns if t.transaction_type == 'OUT']
    audits = [t for t in txns if t.transaction_type == 'AUDIT']

    rentals = list(RentalRecord.objects
                   .filter(product=product)
                   .select_related('rented_out_by', 'returned_by')
                   .order_by('-rent_out_date', '-id'))
    transfers = list(InventoryTransfer.objects
                     .filter(product=product)
                     .select_related('transferred_by')
                     .order_by('-created_at'))
    freezes = list(InventoryFreezeRecord.objects
                   .filter(product=product)
                   .select_related('frozen_by', 'unfrozen_by')
                   .order_by('-frozen_at', '-id'))
    events = list(product.timeline_events.select_related('performed_by').order_by('-created_at'))
    activity = list(ActivityLog.objects
                    .filter(entity_id=str(product.id))
                    .select_related('user')
                    .order_by('-timestamp')[:500])

    latest_txn = txns.first()
    mapping_type, mapping_label, _ = _current_mapping(product)
    detail = _inventory_detail(product)

    return render(request, 'inventory/asset_timeline.html', {
        'product': product,
        'detail': detail,
        'current_warehouse': _latest_store_location(product),
        'current_location': _latest_location(product),
        'current_status': latest_txn.stock_status if latest_txn else '-',
        'is_stocked_out': bool(latest_txn and latest_txn.transaction_type == 'OUT'),
        'is_frozen': _is_frozen(product),
        'mapping_type': mapping_type,
        'mapping_label': mapping_label,
        'on_rent': any(r.status == 'ON_RENT' for r in rentals),
        'events': events,
        'stock_in': stock_in,
        'stock_out': stock_out,
        'audits': audits,
        'rentals': rentals,
        'transfers': transfers,
        'freezes': freezes,
        'activity': activity,
    })


# ════════════════════════════════════════════════════════════
#  MAPPING  (requirement #3)
#
#  Map a standalone inventory item INTO a Server (any category)
#  or a Controller (spares only). Auto-populates warehouse /
#  location / parent reference and keeps mapping + unmapping
#  history in the Activity Trail and the asset timeline.
# ════════════════════════════════════════════════════════════

def _inventory_detail(product):
    """Pull display/identity fields from a product's category record."""
    detail = {'spare_type': '', 'barcode': '', 'part_no': '', 'serial_no': product.serial_no or '', 'obj': None}
    try:
        detail['spare_type'] = (product.category.name or '').upper()
    except Exception:
        pass
    for obj in _related_inventory_objects(product):
        detail['obj'] = obj
        detail['barcode'] = getattr(obj, 'barcode', '') or ''
        detail['part_no'] = getattr(obj, 'part_no', '') or ''
        break
    return detail


def _current_mapping(product):
    """Return ('server'|'controller'|None, label, obj)."""
    from apps.servers.models import ServerComponent
    from apps.categories.models import Spare

    sc = ServerComponent.objects.filter(product=product).select_related('server').first()
    if sc and sc.server:
        s = sc.server
        return 'server', f"{s.machine_no or s.model or 'Server'} [{s.service_tag}]", sc

    spare = Spare.objects.filter(product=product, controller__isnull=False).select_related('controller__product').first()
    if spare and spare.controller:
        c = spare.controller
        return 'controller', f"{c.model or 'Controller'} [{getattr(c.product, 'serial_no', '') or '—'}]", spare
    return None, '', None


@require_permission('mapping')
def mapping_targets(request):
    """AJAX: list servers or controllers for the mapping dropdown (debounced search)."""
    target_type = request.GET.get('type', 'server')
    q = request.GET.get('q', '').strip()
    results = []
    if target_type == 'server':
        from apps.servers.models import Server
        qs = Server.objects.all().order_by('-created_at')
        if q:
            qs = qs.filter(Q(service_tag__icontains=q) | Q(model__icontains=q) | Q(machine_no__icontains=q))
        for s in qs[:20]:
            results.append({'id': s.id, 'label': f"{s.machine_no or s.model or 'Server'} [{s.service_tag}]"})
    else:
        from apps.categories.models import Controller
        qs = Controller.objects.select_related('product').order_by('-id')
        if q:
            qs = qs.filter(Q(product__serial_no__icontains=q) | Q(model__icontains=q) | Q(barcode__icontains=q))
        for c in qs[:20]:
            results.append({'id': c.id, 'label': f"{c.model or 'Controller'} [{getattr(c.product, 'serial_no', '') or '—'}]"})
    return JsonResponse({'results': results})


@require_permission('mapping')
def map_inventory(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    product = get_object_or_404(Product, id=request.POST.get('product_id'))
    target_type = request.POST.get('target_type', 'server')
    target_id = request.POST.get('target_id')
    remarks = request.POST.get('remarks', '').strip()

    if _is_frozen(product):
        return JsonResponse({'success': False, 'error': 'Frozen stock cannot be mapped'})

    existing_type, existing_label, _ = _current_mapping(product)
    if existing_type:
        return JsonResponse({'success': False, 'error': f'Already mapped to {existing_type}: {existing_label}. Unmap first.'})

    detail = _inventory_detail(product)

    with transaction.atomic():
        if target_type == 'server':
            from apps.servers.models import Server, ServerComponent
            server = get_object_or_404(Server, id=target_id)
            ServerComponent.objects.create(
                server=server,
                product=product,
                spare_type=detail['spare_type'],
                part_no=detail['part_no'],
                serial_no=detail['serial_no'],
                barcode=detail['barcode'] or None,
                location=server.location,
                reference_location=server.reference_location,
                parent_child_location=server.service_tag,
                remark=remarks,
            )
            parent_label = f"{server.machine_no or server.model or 'Server'} [{server.service_tag}]"
            parent_warehouse = _latest_store_location(server.product) if server.product else _latest_store_location(product)
            parent_location = server.location or ''
        elif target_type == 'controller':
            from apps.categories.models import Controller, Spare
            controller = get_object_or_404(Controller, id=target_id)
            spare = Spare.objects.filter(product=product).first()
            if not spare:
                return JsonResponse({'success': False, 'error': 'Only spare parts can be mapped to a controller'})
            spare.controller = controller
            spare.location = controller.location or spare.location
            spare.reference_location = controller.reference_location or spare.reference_location
            spare.save(update_fields=['controller', 'location', 'reference_location'])
            parent_label = f"{controller.model or 'Controller'} [{getattr(controller.product, 'serial_no', '') or '—'}]"
            parent_warehouse = _latest_store_location(controller.product) if controller.product else _latest_store_location(product)
            parent_location = controller.location or ''
        else:
            return JsonResponse({'success': False, 'error': 'Invalid target type'})

        # Auto-populate location + warehouse to match the parent.
        if parent_location:
            _set_product_location(product, parent_location)
        current_warehouse = _latest_store_location(product)
        if parent_warehouse and parent_warehouse != current_warehouse:
            last_txn = InventoryTransaction.objects.filter(product=product).order_by('-created_at').first()
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='AUDIT',
                store_location=parent_warehouse,
                stock_status=last_txn.stock_status if last_txn else 'LIVE',
                audited_on=now().date(),
                audited_by=request.user,
                performed_by=request.user,
                audit_remark=f'Mapped into {target_type}: warehouse aligned to {parent_warehouse}',
                audit_result='FOUND',
            )

    log_activity(
        action='MAP',
        module='SERVER' if target_type == 'server' else 'CONTROLLER',
        entity=product.name,
        entity_id=product.id,
        user=request.user,
        barcode=detail['barcode'],
        warehouse=parent_warehouse,
        location=parent_location,
        new_values={'target_type': target_type, 'target_id': target_id, 'parent': parent_label},
        remarks=format_dated_remark(remarks) or f'Mapped into {parent_label}',
    )
    log_timeline(
        product=product, event_type='MAP', user=request.user,
        warehouse=parent_warehouse, location=parent_location,
        remarks=format_dated_remark(remarks) or f'Mapped into {parent_label}',
        details={'target_type': target_type, 'parent': parent_label},
    )
    return JsonResponse({'success': True, 'parent': parent_label})


@require_permission('mapping')
def unmap_inventory(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    product = get_object_or_404(Product, id=request.POST.get('product_id'))
    remarks = request.POST.get('remarks', '').strip()
    mapping_type, label, obj = _current_mapping(product)
    if not mapping_type:
        return JsonResponse({'success': False, 'error': 'This item is not mapped'})

    with transaction.atomic():
        if mapping_type == 'server':
            obj.delete()
        else:  # controller
            obj.controller = None
            obj.save(update_fields=['controller'])

    log_activity(
        action='UNMAP',
        module='SERVER' if mapping_type == 'server' else 'CONTROLLER',
        entity=product.name,
        entity_id=product.id,
        user=request.user,
        old_values={'parent': label, 'target_type': mapping_type},
        remarks=format_dated_remark(remarks) or f'Unmapped from {label}',
    )
    log_timeline(
        product=product, event_type='UNMAP', user=request.user,
        remarks=format_dated_remark(remarks) or f'Unmapped from {label}',
        details={'target_type': mapping_type, 'parent': label},
    )
    return JsonResponse({'success': True})


@require_permission('mapping')
def mapping_status(request):
    """AJAX: current mapping for a product (for the Map modal)."""
    product = get_object_or_404(Product, id=request.GET.get('product_id'))
    mapping_type, label, _ = _current_mapping(product)
    return JsonResponse({'mapped': bool(mapping_type), 'type': mapping_type or '', 'label': label})


# ════════════════════════════════════════════════════════════
#  RENTALS  (rent → return → available again, with history)
# ════════════════════════════════════════════════════════════

@require_permission('rent_return')
def return_rental(request):
    """Return a rented item: closes the rental and brings the asset back into
    active inventory (a fresh Stock In) so it can be sold or rented again."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    rental_id = request.POST.get('rental_id')
    remarks = request.POST.get('remarks', '').strip()
    return_date = _parse_date(request.POST.get('return_date', '')) or date.today()
    # Disposition + warehouse/location, mirroring the Sales Return flow.
    disposition = (request.POST.get('disposition') or 'LIVE').strip().upper()
    if disposition not in {'LIVE', 'FAULTY', 'DAMAGED', 'SCRAP'}:
        disposition = 'LIVE'
    return_warehouse = (request.POST.get('return_warehouse') or '').strip().upper()
    return_location = (request.POST.get('return_location') or '').strip()

    record = get_object_or_404(RentalRecord, id=rental_id)
    if record.status == 'RETURNED':
        return JsonResponse({'success': False, 'error': 'This rental is already returned'})

    product = record.product
    valid_locations = dict(InventoryTransaction.STORE_LOCATION)
    store_location = (
        return_warehouse if return_warehouse in valid_locations
        else (record.store_location or _latest_store_location(product))
    )
    formatted_remarks = _append_remark_to_asset(product, remarks, return_date)

    with transaction.atomic():
        record.status = 'RETURNED'
        record.actual_return_date = return_date
        record.returned_by = request.user if request.user.is_authenticated else None
        if formatted_remarks:
            record.remarks = (record.remarks + ' | ' if record.remarks else '') + f'Return: {formatted_remarks}'
        record.save(update_fields=['status', 'actual_return_date', 'returned_by', 'remarks'])

        if disposition == 'SCRAP':
            # Returned but scrapped → stays out of active stock.
            _create_scrap_transaction(product, remarks, request.user, scrapped_on=return_date,
                                      scrap_location=store_location)
        else:
            # Bring the asset back in stock (LIVE / FAULTY / DAMAGED) → it reappears
            # in active lists and is available to sell or rent again.
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='IN',
                store_location=store_location,
                stock_status=disposition,
                stock_in_date=return_date,
                performed_by=request.user if request.user.is_authenticated else None,
            )
        if return_location:
            _set_product_location(product, return_location)

    log_activity(
        action='RENT_RETURN',
        module='INVENTORY',
        entity=product.name,
        entity_id=product.id,
        user=request.user,
        warehouse=store_location,
        location=return_location or _latest_location(product),
        old_values={'status': 'ON_RENT', 'client': record.client_name},
        new_values={'status': 'RETURNED', 'return_date': str(return_date),
                    'disposition': disposition, 'available': disposition != 'SCRAP'},
        remarks=formatted_remarks or 'Rental returned — back in stock',
    )
    log_timeline(
        product=product, event_type='IN', user=request.user,
        warehouse=store_location, remarks=formatted_remarks or 'Rental returned — back in stock',
        details={'rental_id': record.id, 'client': record.client_name},
    )
    return JsonResponse({'success': True})


@require_permission('rent_return')
def rental_list(request):
    """Rental register: current rentals (On Rent) and full rent history.
    Default view is On Rent; filter to Returned/All for history."""
    base = RentalRecord.objects.select_related('product', 'rented_out_by', 'returned_by')
    total_on_rent = base.filter(status='ON_RENT').count()
    total_all = base.count()

    status = request.GET.get('status', 'ON_RENT').strip()
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    records = base.order_by('-rent_out_date', '-id')
    if status in ('ON_RENT', 'RETURNED'):
        records = records.filter(status=status)
    if q:
        records = records.filter(
            Q(product__name__icontains=q) |
            Q(product__serial_no__icontains=q) |
            Q(client_name__icontains=q) |
            Q(invoice_no__icontains=q) |
            Q(olf_dc_number__icontains=q)
        )
    if date_from:
        records = records.filter(rent_out_date__gte=date_from)
    if date_to:
        records = records.filter(rent_out_date__lte=date_to)

    records = list(records)
    return render(request, 'inventory/rental_list.html', {
        'records': records,
        'status': status,
        'total_on_rent': total_on_rent,
        'total_all': total_all,
        'result_count': len(records),
        'filters': {'q': q, 'date_from': date_from, 'date_to': date_to},
    })


@require_permission('mapping')
def mapping_history(request):
    logs = ActivityLog.objects.filter(action__in=['MAP', 'UNMAP']).select_related('user').order_by('-timestamp')
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if q:
        logs = logs.filter(Q(entity__icontains=q) | Q(barcode__icontains=q) | Q(remarks__icontains=q))
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)
    total_count = ActivityLog.objects.filter(action__in=['MAP', 'UNMAP']).count()
    logs = list(logs[:1000])
    rows = []
    for log in logs:
        parent = ''
        if isinstance(log.new_values, dict):
            parent = log.new_values.get('parent', '')
        if not parent and isinstance(log.old_values, dict):
            parent = log.old_values.get('parent', '')
        rows.append({'log': log, 'parent': parent})
    return render(request, 'inventory/mapping_history.html', {
        'rows': rows,
        'total_count': total_count,
        'result_count': len(rows),
        'filters': {'q': q, 'date_from': date_from, 'date_to': date_to},
    })


# ════════════════════════════════════════════════════════════
#  AUDIT EXCEL RECONCILIATION  (requirement #8)
#
#  Upload an Excel of physical-audit barcodes. For each row:
#    • barcode found  → update Warehouse / Location / Rack / Bin,
#                       log old → new in the Activity Trail + timeline
#    • barcode missing → auto-create an Audit Finding
#                        (Found During Audit / Missing Inventory Record)
#  Shows an import summary and a downloadable error report.
# ════════════════════════════════════════════════════════════

RECON_COLUMNS = {
    'Barcode': 'barcode',
    'Warehouse': 'warehouse',
    'Location': 'location',
    'Rack': 'rack',
    'Bin': 'bin',
}


def _find_inventory_by_barcode(barcode):
    """Return (obj, product) for the first inventory record matching *barcode*,
    searched across every category + server model. obj exposes .location."""
    from apps.categories.models import (
        Card, CPU, HardDisk, Memory, NetworkingSpare, RailKit, SFP, Spare, Controller,
    )
    from apps.servers.models import Server, ServerComponent

    models = (Spare, Card, CPU, Memory, SFP, RailKit, HardDisk, Controller,
              NetworkingSpare, Server, ServerComponent)
    for model in models:
        obj = model.objects.filter(barcode__iexact=barcode).select_related('product').first()
        if obj and getattr(obj, 'product', None):
            return obj, obj.product
    return None, None


def _reconcile_row(row, user):
    """Apply one reconciliation row. Returns ('updated'|'matched'|'unmatched', detail)."""
    barcode = (row.get('barcode') or '').strip()
    if not barcode:
        raise ValueError('Missing barcode')

    new_warehouse = (row.get('warehouse') or '').strip().upper()
    new_location = (row.get('location') or '').strip()
    rack = (row.get('rack') or '').strip()
    bin_ = (row.get('bin') or '').strip()

    obj, product = _find_inventory_by_barcode(barcode)

    if not product:
        # Missing inventory record → auto-create an Audit Finding.
        AuditFinding.objects.create(
            audit_date=now().date(),
            remarks=f'Found During Audit | Missing Inventory Record | Barcode: {barcode} '
                    f'| Warehouse: {new_warehouse or "-"} | Location: {new_location or "-"}',
            person_involved=getattr(user, 'username', '') or '',
            created_by=user if getattr(user, 'is_authenticated', False) else None,
        )
        return 'unmatched', f'No record for barcode {barcode} — Audit Finding created'

    old_warehouse = _latest_store_location(product)
    old_location = getattr(obj, 'location', '') or ''

    # Build the reference (rack/bin) string.
    ref_bits = []
    if rack:
        ref_bits.append(f'Rack: {rack}')
    if bin_:
        ref_bits.append(f'Bin: {bin_}')
    new_reference = ' / '.join(ref_bits)

    changed = False
    with transaction.atomic():
        if new_location and new_location != old_location:
            obj.location = new_location
            changed = True
        if new_reference and hasattr(obj, 'reference_location'):
            obj.reference_location = new_reference
            changed = True
        if changed:
            update_fields = ['location']
            if hasattr(obj, 'reference_location'):
                update_fields.append('reference_location')
            obj.save(update_fields=update_fields)

        warehouse_changed = bool(new_warehouse) and new_warehouse != old_warehouse
        if warehouse_changed:
            valid = dict(InventoryTransaction.STORE_LOCATION)
            if new_warehouse not in valid:
                raise ValueError(f'Invalid warehouse: {new_warehouse}')
            last_txn = InventoryTransaction.objects.filter(product=product).order_by('-created_at').first()
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='AUDIT',
                store_location=new_warehouse,
                stock_status=last_txn.stock_status if last_txn else 'LIVE',
                audited_on=now().date(),
                audited_by=user if getattr(user, 'is_authenticated', False) else None,
                performed_by=user if getattr(user, 'is_authenticated', False) else None,
                audit_remark=f'Reconciliation: warehouse {old_warehouse} → {new_warehouse}',
                audit_result='FOUND',
            )
            changed = True

    if changed:
        log_activity(
            action='RECONCILIATION_UPDATE',
            module='AUDIT',
            entity=product.name,
            entity_id=product.id,
            user=user,
            barcode=barcode,
            warehouse=new_warehouse or old_warehouse,
            location=new_location or old_location,
            old_values={'warehouse': old_warehouse, 'location': old_location},
            new_values={'warehouse': new_warehouse or old_warehouse,
                        'location': new_location or old_location,
                        'reference': new_reference},
            remarks='Updated via audit reconciliation',
        )
        log_timeline(
            product=product, event_type='AUDIT', user=user,
            warehouse=new_warehouse or old_warehouse, location=new_location or old_location,
            remarks='Audit reconciliation update',
        )
        return 'updated', f'{barcode}: warehouse {old_warehouse}→{new_warehouse or old_warehouse}, location {old_location or "-"}→{new_location or old_location or "-"}'

    return 'matched', f'{barcode}: found, no change'


@require_permission('reconciliation')
def audit_reconciliation(request):
    if request.method == 'POST':
        from openpyxl import load_workbook
        upload = request.FILES.get('file')
        if not upload:
            return render(request, 'inventory/reconciliation.html', {
                'error': 'Please choose an Excel file to upload.',
            })

        summary = {'total': 0, 'updated': 0, 'matched': 0, 'unmatched': 0, 'failed': 0}
        errors = []
        log_activity(action='RECONCILIATION_STARTED', module='AUDIT', entity='Reconciliation',
                     user=request.user, remarks=f'File: {upload.name}')
        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [str(c.value).strip() if c.value is not None else '' for c in header_cells]
            field_map = {src: tgt for src, tgt in RECON_COLUMNS.items() if src in headers}
            if 'Barcode' not in headers:
                return render(request, 'inventory/reconciliation.html', {
                    'error': 'The file must contain a "Barcode" column. '
                             'Expected columns: ' + ', '.join(RECON_COLUMNS.keys()),
                })

            for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(v is not None and str(v).strip() for v in values):
                    continue
                summary['total'] += 1
                raw = dict(zip(headers, values))
                row = {tgt: (str(raw.get(src)).strip() if raw.get(src) is not None else '')
                       for src, tgt in field_map.items()}
                try:
                    outcome, _detail = _reconcile_row(row, request.user)
                    summary[outcome] += 1
                except Exception as exc:  # noqa: BLE001 - per-row isolation
                    summary['failed'] += 1
                    errors.append({'row': idx, 'barcode': row.get('barcode', ''), 'error': str(exc)})
        except Exception as exc:  # noqa: BLE001
            return render(request, 'inventory/reconciliation.html', {
                'error': f'Could not read the file: {exc}',
            })

        request.session['recon_errors'] = errors
        log_activity(action='RECONCILIATION_COMPLETED', module='AUDIT', entity='Reconciliation',
                     user=request.user,
                     new_values=summary, remarks=f'File: {upload.name}')
        return render(request, 'inventory/reconciliation.html', {
            'summary': summary,
            'errors': errors,
            'completed': True,
            'file_name': upload.name,
        })

    return render(request, 'inventory/reconciliation.html', {
        'columns': list(RECON_COLUMNS.keys()),
    })


@require_permission('reconciliation')
def reconciliation_error_report(request):
    import csv
    from django.http import HttpResponse
    errors = request.session.get('recon_errors', [])
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reconciliation-errors.csv"'
    writer = csv.writer(response)
    writer.writerow(['Row', 'Barcode', 'Error'])
    for e in errors:
        writer.writerow([e.get('row', ''), e.get('barcode', ''), e.get('error', '')])
    return response
