from datetime import date
from urllib import request
from django.db.models import OuterRef, Q, Subquery

from django.http import JsonResponse
from django.shortcuts import render, redirect
from apps.categories.models import CPU, Spare, Card, Controller, NetworkingSpare, Memory, SFP, RailKit, HardDisk
from apps.core.services import create_product_and_railkit, create_controller_with_components, create_product_and_card, create_product_and_cpu, create_product_and_memory, create_product_and_sfp, create_product_and_spare, create_product_and_harddisk
from apps.core.models import Product, SpareCategory, Brand
from django.db.models import OuterRef, Subquery, Value, CharField
from django.db.models.functions import Coalesce, Concat
from django.contrib.postgres.aggregates import StringAgg
from apps.servers.models import ServerComponent

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views.decorators.http import require_POST
from apps.core.permissions import require_any_permission

from apps.inventory.models import InventoryFreezeRecord, InventoryTransaction
from apps.core.permissions import has_permission
from apps.core.activity import append_dated_remark
from django.http import HttpResponseForbidden


def _exclude_out_or_frozen(qs):
    latest_freeze = InventoryFreezeRecord.objects.filter(
        product=OuterRef('product')
    ).order_by('-frozen_at', '-id')
    return qs.annotate(
        latest_freeze_status=Subquery(latest_freeze.values('status')[:1])
    ).exclude(latest_type='OUT').filter(
        Q(latest_freeze_status__isnull=True) | ~Q(latest_freeze_status='FROZEN')
    )


def _mark_latest_stock_in_user(product, user):
    if not product or not getattr(user, 'is_authenticated', False):
        return
    txn = InventoryTransaction.objects.filter(
        product=product,
        transaction_type='IN',
    ).order_by('-created_at').first()
    if txn:
        txn.performed_by = user
        txn.save(update_fields=['performed_by'])

def add_spare(request):
    if request.method == "POST":

        try:
            data_map = {
                'category': request.POST.getlist('category[]'),
                'brand': request.POST.getlist('brand[]'),
                'model': request.POST.getlist('model[]'),
                'part_no': request.POST.getlist('part_no[]'),
                'alt_part_no': request.POST.getlist('alt_part_no[]'),
                'serial_no': request.POST.getlist('serial_no[]'),
                'alt_serial_no': request.POST.getlist('alt_serial_no[]'),
                'specs': request.POST.getlist('specs[]'),
                'qty': request.POST.getlist('qty[]'),
                'barcode': request.POST.getlist('barcode[]'),
                'location': request.POST.getlist('location[]'),
                'reference_location': request.POST.getlist('reference_location[]'),
                'remark': request.POST.getlist('remark[]'),
                'store_location': request.POST.getlist('store_location[]'),
                'stock_status': request.POST.getlist('stock_status[]'),
            }

            total = len(data_map['serial_no'])

            for i in range(total):
                if not data_map['serial_no'][i]:
                    continue

                row = {k: v[i] for k, v in data_map.items()}

                product = create_product_and_spare(row)
                _mark_latest_stock_in_user(product, request.user)

            messages.success(request, "Spare added successfully")

        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'spare/add_spare.html', {
        'categories': SpareCategory.objects.all(),
        'brands': Brand.objects.all()
    })

def check_serial(request):
    serial = request.GET.get('serial_no')

    exists = Product.objects.filter(serial_no=serial).exists()

    return JsonResponse({'exists': exists})



def spare_list(request):
    from apps.categories.models import Spare
    from apps.inventory.models import InventoryTransaction
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    spares = Spare.objects.select_related(
        'product', 'product__category',
        'controller', 'controller__product',
    ).annotate(
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date = Subquery(latest_audit.values('audited_on')[:1]),
        # server membership
        server_label    = Subquery(
            ServerComponent.objects.filter(
                product=OuterRef('product')
            ).values('server__machine_no')[:1]
        ),
        server_service_tag = Subquery(
            ServerComponent.objects.filter(
                product=OuterRef('product')
            ).values('server__service_tag')[:1]
        ),
        server_model_label = Subquery(
            ServerComponent.objects.filter(
                product=OuterRef('product')
            ).values('server__model')[:1]
        ),
        server_pk = Subquery(
            ServerComponent.objects.filter(
                product=OuterRef('product')
            ).values('server__id')[:1]
        ),
    )
    spares = _exclude_out_or_frozen(spares)
 
    return render(request, 'spare/list.html', {'spares': spares})
 
 
# # NEW VIEW — components API used by controller list page
# def controller_components(request, controller_id):
#     from apps.categories.models import Controller
 
#     try:
#         ctrl = _exclude_out_or_frozen(Controller.objects.get(id=controller_id)
#     except Controller.DoesNotExist:
#         return JsonResponse({'components': []})
 
#     latest_txn = InventoryTransaction.objects.filter(
#         product=OuterRef('product')
#     ).order_by('-created_at')
 
#     components = ctrl.components.select_related(
#         'product', 'product__category', 'brand'
#     ).annotate(
#         latest_location=Subquery(latest_txn.values('store_location')[:1]),
#         latest_status=Subquery(latest_txn.values('stock_status')[:1]),
#     )
 
#     data = []
#     for comp in components:
#         data.append({
#             'category':  comp.product.category.name,
#             'brand':     str(comp.brand) if comp.brand else '',
#             'model':     comp.model or '',
#             'part_no':   comp.part_no or '',
#             'serial_no': comp.product.serial_no,
#             'specs':     comp.specs or '',
#             'barcode':   comp.barcode or '',
#             'store':     comp.latest_location or '-',
#             'status':    comp.latest_status or '-',
#         })
 
#     return JsonResponse({'components': data})

# 🔥 INLINE UPDATE (AJAX)
def _update_remark_field(model_cls, obj_id, value):
    obj = model_cls.objects.get(id=obj_id)
    current = getattr(obj, 'remark', '')
    setattr(obj, 'remark', append_dated_remark(current, value))
    obj.save(update_fields=['remark'])
    return obj


def update_spare_field(request):
    if request.method == "POST":
        model_name = (request.POST.get('model') or 'spare').strip().lower()
        spare_id = request.POST.get('id')
        field = request.POST.get('field')
        value = request.POST.get('value')

        if field != 'remark':
            return JsonResponse({'success': False, 'error': 'Only remarks can be updated from a list'})

        model_map = {
            'spare': Spare,
            'card': Card,
            'cpu': CPU,
            'controller': Controller,
            'memory': Memory,
            'sfp': SFP,
            'railkit': RailKit,
            'harddisk': HardDisk,
        }
        model_cls = model_map.get(model_name, Spare)

        try:
            _update_remark_field(model_cls, spare_id, value)
        except model_cls.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Record not found'})

        return JsonResponse({'success': True})

    return JsonResponse({'success': False})


@require_any_permission('sold_view', 'reports')
def spare_out_report(request):
    selected_status = request.GET.get('status', '').strip()
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')

    all_out_spares = Spare.objects.select_related('product').annotate(
        latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location=Subquery(latest_txn.values('store_location')[:1]),
        latest_status=Subquery(latest_txn.values('stock_status')[:1]),
        latest_client=Subquery(latest_txn.values('client_name')[:1]),
        latest_invoice=Subquery(latest_txn.values('invoice_no')[:1]),
        latest_olf_dc=Subquery(latest_txn.values('olf_dc_number')[:1]),
        latest_out_date=Subquery(latest_txn.values('stock_out_date')[:1]),
    ).filter(
        latest_type='OUT'   # 🔥 ONLY STOCKED OUT
    )

    available_statuses = sorted([
        s for s in all_out_spares.values_list('latest_status', flat=True).distinct() if s
    ])

    total_count = all_out_spares.count()

    spares = all_out_spares
    if selected_status:
        spares = spares.filter(latest_status=selected_status)
    if date_from:
        spares = spares.filter(latest_out_date__gte=date_from)
    if date_to:
        spares = spares.filter(latest_out_date__lte=date_to)
    if q:
        spares = spares.filter(
            Q(product__name__icontains=q) |
            Q(product__serial_no__icontains=q) |
            Q(part_no__icontains=q) |
            Q(barcode__icontains=q) |
            Q(latest_client__icontains=q) |
            Q(latest_invoice__icontains=q) |
            Q(latest_olf_dc__icontains=q)
        )

    spares = list(spares)
    return render(request, 'spare/spare_out_report.html', {
        'spares': spares,
        'available_statuses': available_statuses,
        'selected_status': selected_status,
        'total_count': total_count,
        'result_count': len(spares),
        'filters': {'q': q, 'date_from': date_from, 'date_to': date_to},
        'can_stock_return': has_permission(request.user, 'stock_return'),
    })


def add_card(request):
    if request.method == "POST":

        data_map = {
            'brand': request.POST.getlist('brand[]'),
            'oem': request.POST.getlist('oem[]'),
            'brand_model_no': request.POST.getlist('brand_model_no[]'),
            'interface': request.POST.getlist('interface[]'),
            'part_no': request.POST.getlist('part_no[]'),
            'alt_part_no': request.POST.getlist('alt_part_no[]'),
            'serial_no': request.POST.getlist('brand_serial_no_1[]'),
            'brand_serial_no_1': request.POST.getlist('brand_serial_no_1[]'),
            'capacity': request.POST.getlist('capacity[]'),
            'port': request.POST.getlist('port[]'),
            'barcode': request.POST.getlist('barcode[]'),
            'location': request.POST.getlist('location[]'),
            'reference_location': request.POST.getlist('reference_location[]'),
            'remark': request.POST.getlist('remark[]'),
            'store_location': request.POST.getlist('store_location[]'),
            'stock_status': request.POST.getlist('stock_status[]'),
        }

        total = len(data_map['serial_no'])

        for i in range(total):
            if not data_map['serial_no'][i]:
                continue

            row = {k: v[i] for k, v in data_map.items()}

            product = create_product_and_card(row)
            _mark_latest_stock_in_user(product, request.user)

    return render(request, 'card/add_card.html', {'brands': Brand.objects.all()})



def card_list(request):

    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')

    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')

    cards = Card.objects.select_related('product').annotate(
        latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location=Subquery(latest_txn.values('store_location')[:1]),
        latest_status=Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date=Subquery(latest_audit.values('audited_on')[:1]),
    )
    cards = _exclude_out_or_frozen(cards)

    return render(request, 'card/list.html', {
        'cards': cards
    })


def add_cpu(request):
    if request.method == "POST":
        try:
            data_map = {
                'brand': request.POST.getlist('brand[]'),
                'model': request.POST.getlist('model[]'),
                'part_no': request.POST.getlist('part_no[]'),
                'serial_no': request.POST.getlist('part_no[]'),
                'no_of_cores': request.POST.getlist('no_of_cores[]'),
                'no_of_threads': request.POST.getlist('no_of_threads[]'),
                'ghz': request.POST.getlist('ghz[]'),
                'frequency': request.POST.getlist('frequency[]'),
                'cache': request.POST.getlist('cache[]'),
                'barcode': request.POST.getlist('barcode[]'),
                'location': request.POST.getlist('location[]'),
                'reference_location': request.POST.getlist('reference_location[]'),
                'remark': request.POST.getlist('remark[]'),
                'store_location': request.POST.getlist('store_location[]'),
                'stock_status': request.POST.getlist('stock_status[]'),
            }

            total = len(data_map['serial_no'])

            for i in range(total):
                serial = data_map['serial_no'][i].strip()
                if not serial:
                    continue

                row = {}
                for key, value_list in data_map.items():
                    try:
                        row[key] = value_list[i]
                    except IndexError:
                        row[key] = None

                if not row.get('store_location'):
                    row['store_location'] = 'WH1'
                if not row.get('stock_status'):
                    row['stock_status'] = 'LIVE'

                product = create_product_and_cpu(row)
                _mark_latest_stock_in_user(product, request.user)

            messages.success(request, "PROCESSOR added successfully")
            return redirect('add_cpu')

        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'cpu/add_cpu.html', {
        'brands': Brand.objects.all()
    })


def cpu_list(request):
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    cpus = CPU.objects.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location=Subquery(latest_txn.values('store_location')[:1]),
        latest_status=Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date=Subquery(latest_audit.values('audited_on')[:1]),
    )
    cpus = _exclude_out_or_frozen(cpus)
 
    return render(request, 'cpu/cpu_list.html', {
        'cpus': cpus
    })
 
 
def update_cpu_field(request):
    """Inline edit handler for CPU fields: location, reference_location, remark"""
 
    ALLOWED_FIELDS = ['remark']
 
    if request.method == 'POST':
        cpu_id = request.POST.get('id')
        field = request.POST.get('field')
        value = request.POST.get('value')
 
        if field not in ALLOWED_FIELDS:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            cpu = CPU.objects.get(id=cpu_id)
            setattr(cpu, field, append_dated_remark(getattr(cpu, field), value))
            cpu.save()
            return JsonResponse({'success': True})
        except CPU.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'CPU not found'})
 
    return JsonResponse({'success': False})



def add_controller(request):
    if request.method == 'POST':
        try:
            controller_data = {
                'product_name':         request.POST.get('controller_product_name', ''),
                'brand':                request.POST.get('cabinet_brand', ''),
                'model':                request.POST.get('cabinet_model', ''),
                'part_no':              request.POST.get('cabinet_part_no', ''),
                'alt_part_no':          request.POST.get('cabinet_alt_part_no', ''),
                'serial_no':            request.POST.get('cabinet_serial_no', ''),
                'alt_serial_no':        request.POST.get('cabinet_alt_serial_no', ''),
                'specs':                request.POST.get('cabinet_specs', ''),
                'qty':                  request.POST.get('cabinet_qty', 1),
                'barcode':              request.POST.get('cabinet_barcode', ''),
                'location':             request.POST.get('cabinet_location', ''),
                'reference_location':   request.POST.get('cabinet_reference_location', ''),
                'parent_child_location':request.POST.get('cabinet_parent_child_location', ''),
                'remark':               request.POST.get('cabinet_remark', ''),
                'store_location':       request.POST.get('cabinet_store_location', 'WH1'),
                'stock_status':         request.POST.get('cabinet_stock_status', 'LIVE'),
            }
 
            # --- COMPONENTS (multi-row) ---
            fields = [
                'spare_type', 'product_name', 'brand', 'model', 'part_no',
                'alt_part_no', 'serial_no', 'alt_serial_no',
                'specs', 'qty', 'barcode', 'remark',
            ]
 
            lists = {f: request.POST.getlist(f'comp_{f}[]') for f in fields}
            total = len(lists['serial_no'])
 
            components = []
            for i in range(total):
                row = {}
                for f in fields:
                    try:
                        row[f] = lists[f][i]
                    except IndexError:
                        row[f] = ''
                components.append(row)
 
            controller = create_controller_with_components(controller_data, components)
            _mark_latest_stock_in_user(controller.product, request.user)
            for component in controller.components.select_related('product'):
                _mark_latest_stock_in_user(component.product, request.user)
 
            messages.success(request, 'Controller created successfully.')
            return redirect('add_controller')
 
        except Exception as e:
            messages.error(request, str(e))
 
    return render(request, 'controller/add_controller.html', {
        'brands': Brand.objects.all(),
        'spare_categories': SpareCategory.objects.all(),
    })
 
 
def controller_list(request):
    from apps.categories.models import Controller
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    controllers = Controller.objects.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date = Subquery(latest_audit.values('audited_on')[:1]),
        component_search = StringAgg(
            Concat(
                Coalesce('components__product__name', Value('')),
                Value(' '),
                Coalesce('components__part_no', Value('')),
                Value(' '),
                Coalesce('components__alt_part_no', Value('')),
                Value(' '),
                Coalesce('components__product__serial_no', Value('')),
                Value(' '),
                Coalesce('components__barcode', Value('')),
                output_field=CharField(),
            ),
            delimiter=' ',
            distinct=True,
        ),
    )
    controllers = _exclude_out_or_frozen(controllers)
 
    return render(request, 'controller/controller_list.html', {
        'controllers':      controllers,
        'spare_categories': SpareCategory.objects.all(),   # for add-component modal
        'brands':           Brand.objects.all(),           # for add-component modal
    })
 
 
def update_controller_field(request):
    from apps.categories.models import Controller
 
    ALLOWED_FIELDS = ['remark']
 
    if request.method == 'POST':
        ctrl_id = request.POST.get('id')
        field   = request.POST.get('field')
        value   = request.POST.get('value')
 
        if field not in ALLOWED_FIELDS:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            ctrl = Controller.objects.get(id=ctrl_id)
            setattr(ctrl, field, append_dated_remark(getattr(ctrl, field), value))
            ctrl.save()
            return JsonResponse({'success': True})
        except Controller.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
 
    return JsonResponse({'success': False})


def add_controller_component(request):
    """
    AJAX endpoint — adds one or more spare components to an
    existing Controller. Called from the controller list modal.
    """
    from apps.categories.models import Controller, Spare
 
    if request.method == 'POST':
        try:
            controller_id = request.POST.get('controller_id')
            controller    = Controller.objects.select_related(
                'product'
            ).get(id=controller_id)
 
            fields = [
                'spare_type', 'product_name', 'brand', 'model', 'part_no',
                'alt_part_no', 'serial_no', 'alt_serial_no',
                'specs', 'qty', 'barcode', 'remark',
            ]
 
            lists = {f: request.POST.getlist(f'comp_{f}[]') for f in fields}
            total = len(lists['serial_no'])
 
            added = 0
 
            for i in range(total):
                serial = lists['serial_no'][i].strip().upper()
                if not serial:
                    continue
 
                # duplicate check
                if Product.objects.filter(serial_no=serial).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'Serial already exists: {serial}'
                    })
 
                row = {}
                for f in fields:
                    try:
                        row[f] = lists[f][i]
                    except IndexError:
                        row[f] = ''
 
                comp_category_name = (row.get('spare_type') or '').strip().upper()
                comp_category, _   = SpareCategory.objects.get_or_create(
                    name=comp_category_name
                )
 
                comp_brand_name = row.get('brand', '').strip().upper()
                comp_brand = None
                if comp_brand_name:
                    comp_brand, _ = Brand.objects.get_or_create(name=comp_brand_name)
 
                comp_name = row.get('product_name') or (
                    f'CONTROLLER - '
                    f"{comp_category_name} - "
                    f"{row.get('model', '')} - "
                    f"{comp_brand_name}"
                ).strip(' -')
 
                comp_product = Product.objects.create(
                    category  = comp_category,
                    serial_no = serial,
                    part_no   = row.get('part_no'),
                    brand     = comp_brand,
                    model     = row.get('model'),
                    name      = comp_name
                )
 
                Spare.objects.create(
                    product            = comp_product,
                    controller         = controller,
                    brand              = comp_brand,
                    model              = row.get('model'),
                    part_no            = row.get('part_no'),
                    alt_part_no        = row.get('alt_part_no'),
                    alt_serial_no      = row.get('alt_serial_no'),
                    specs              = row.get('specs'),
                    qty                = row.get('qty') or 1,
                    barcode            = row.get('barcode'),
                    # inherit controller location
                    location           = controller.location,
                    reference_location = controller.reference_location,
                    remark             = row.get('remark'),
                )
 
                # auto stock IN — inherit controller's last transaction location/status
                last_ctrl_txn = InventoryTransaction.objects.filter(
                    product=controller.product
                ).order_by('-created_at').first()
 
                from apps.core.services import _create_stock_in_transaction
                _create_stock_in_transaction(
                    product=comp_product,
                    store_location=last_ctrl_txn.store_location if last_ctrl_txn else 'WH1',
                    stock_status=last_ctrl_txn.stock_status if last_ctrl_txn else 'LIVE',
                    user=request.user,
                    remarks='Controller component added manually',
                )
 
                added += 1
 
            return JsonResponse({'success': True, 'added': added})
 
        except Controller.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Controller not found'})
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'error': str(e),
                'trace': traceback.format_exc()
            })
 
    return JsonResponse({'success': False})
 
 
def controller_components(request, controller_id):
    """
    Returns all components for a controller with their latest
    transaction status — OUT items are flagged.
    """
    from apps.categories.models import Controller
 
    try:
        ctrl = Controller.objects.get(id=controller_id)
    except Controller.DoesNotExist:
        return JsonResponse({'components': []})
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    components = ctrl.components.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_client   = Subquery(latest_txn.values('client_name')[:1]),
        latest_invoice  = Subquery(latest_txn.values('invoice_no')[:1]),
        latest_out_date = Subquery(latest_txn.values('stock_out_date')[:1]),
    )
 
    data = []
    for comp in components:
        is_out = (comp.latest_type == 'OUT')
        data.append({
            'id':          comp.id,
            'product_id':  comp.product.id,
            'category':    comp.product.category.name,
            'brand':       str(comp.brand) if comp.brand else '',
            'model':       comp.model    or '',
            'part_no':     comp.part_no  or '',
            'serial_no':   comp.product.serial_no,
            'specs':       comp.specs    or '',
            'barcode':     comp.barcode  or '',
            'store':       comp.latest_location or '-',
            'status':      comp.latest_status   or '-',
            'is_out':      is_out,
            'client':      comp.latest_client   or '',
            'invoice':     comp.latest_invoice  or '',
            'out_date':    str(comp.latest_out_date) if comp.latest_out_date else '',
        })
 
    return JsonResponse({'components': data})


def add_memory(request):
    if request.method == 'POST':
        try:
            data_map = {
                'brand':              request.POST.getlist('brand[]'),
                'oem':                request.POST.getlist('oem[]'),
                'model':              request.POST.getlist('model[]'),
                'part_no_1':          request.POST.getlist('part_no_1[]'),
                'part_no_2':          request.POST.getlist('part_no_2[]'),
                'part_no_3':          request.POST.getlist('part_no_3[]'),
                'size':               request.POST.getlist('size[]'),
                'ram_type':           request.POST.getlist('ram_type[]'),
                'ddr_version':        request.POST.getlist('ddr_version[]'),
                'frequency':          request.POST.getlist('frequency[]'),
                'rank':               request.POST.getlist('rank[]'),
                'serial_no':          request.POST.getlist('serial_no[]'),
                'qty':                request.POST.getlist('qty[]'),
                'barcode':            request.POST.getlist('barcode[]'),
                'location':           request.POST.getlist('location[]'),
                'reference_location': request.POST.getlist('reference_location[]'),
                'remark':             request.POST.getlist('remark[]'),
                'store_location':     request.POST.getlist('store_location[]'),
                'stock_status':       request.POST.getlist('stock_status[]'),
            }
 
            total = len(data_map['serial_no'])
 
            for i in range(total):
                serial = data_map['serial_no'][i].strip()
                if not serial:
                    continue
 
                row = {}
                for key, value_list in data_map.items():
                    try:
                        row[key] = value_list[i]
                    except IndexError:
                        row[key] = None
 
                if not row.get('store_location'):
                    row['store_location'] = 'WH1'
                if not row.get('stock_status'):
                    row['stock_status'] = 'LIVE'
 
                product = create_product_and_memory(row)
                _mark_latest_stock_in_user(product, request.user)
 
            messages.success(request, 'Memory added successfully.')
            return redirect('add_memory')
 
        except Exception as e:
            messages.error(request, str(e))
 
    return render(request, 'memory/add_memory.html', {
        'brands': Brand.objects.all(),
    })
 
 
def memory_list(request):
    from apps.categories.models import Memory
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    memories = Memory.objects.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_type      = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location  = Subquery(latest_txn.values('store_location')[:1]),
        latest_status    = Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date  = Subquery(latest_audit.values('audited_on')[:1]),
    )
    memories = _exclude_out_or_frozen(memories)
 
    return render(request, 'memory/memory_list.html', {'memories': memories})
 
 
def update_memory_field(request):
    from apps.categories.models import Memory
 
    ALLOWED_FIELDS = ['remark']
 
    if request.method == 'POST':
        mem_id = request.POST.get('id')
        field  = request.POST.get('field')
        value  = request.POST.get('value')
 
        if field not in ALLOWED_FIELDS:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            mem = Memory.objects.get(id=mem_id)
            setattr(mem, field, append_dated_remark(getattr(mem, field), value))
            mem.save()
            return JsonResponse({'success': True})
        except Memory.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
 
    return JsonResponse({'success': False})



def add_sfp(request):
    if request.method == 'POST':
        try:
            data_map = {
                'brand':              request.POST.getlist('brand[]'),
                'model':              request.POST.getlist('model[]'),
                'part_no':            request.POST.getlist('part_no[]'),
                'description':        request.POST.getlist('description[]'),
                'fibre_type':         request.POST.getlist('fibre_type[]'),
                'data_rate':          request.POST.getlist('data_rate[]'),
                'serial_no':          request.POST.getlist('serial_no[]'),
                'barcode':            request.POST.getlist('barcode[]'),
                'location':           request.POST.getlist('location[]'),
                'reference_location': request.POST.getlist('reference_location[]'),
                'remark':             request.POST.getlist('remark[]'),
                'store_location':     request.POST.getlist('store_location[]'),
                'stock_status':       request.POST.getlist('stock_status[]'),
            }
 
            total = len(data_map['serial_no'])
 
            for i in range(total):
                serial = data_map['serial_no'][i].strip()
                if not serial:
                    continue
 
                row = {}
                for key, value_list in data_map.items():
                    try:
                        row[key] = value_list[i]
                    except IndexError:
                        row[key] = None
 
                if not row.get('store_location'):
                    row['store_location'] = 'WH1'
                if not row.get('stock_status'):
                    row['stock_status'] = 'LIVE'
 
                product = create_product_and_sfp(row)
                _mark_latest_stock_in_user(product, request.user)
 
            messages.success(request, 'SFP added successfully.')
            return redirect('add_sfp')
 
        except Exception as e:
            messages.error(request, str(e))
 
    return render(request, 'sfp/add_sfp.html', {
        'brands': Brand.objects.all(),
    })
 
 
def sfp_list(request):
    from apps.categories.models import SFP
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    sfps = SFP.objects.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date = Subquery(latest_audit.values('audited_on')[:1]),
    )
    sfps = _exclude_out_or_frozen(sfps)
 
    return render(request, 'sfp/sfp_list.html', {'sfps': sfps})
 
 
def update_sfp_field(request):
    from apps.categories.models import SFP
 
    ALLOWED_FIELDS = ['remark']
 
    if request.method == 'POST':
        sfp_id = request.POST.get('id')
        field  = request.POST.get('field')
        value  = request.POST.get('value')
 
        if field not in ALLOWED_FIELDS:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            sfp = SFP.objects.get(id=sfp_id)
            setattr(sfp, field, append_dated_remark(getattr(sfp, field), value))
            sfp.save()
            return JsonResponse({'success': True})
        except SFP.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
 
    return JsonResponse({'success': False})



def add_railkit(request):
    if request.method == 'POST':
        try:
            data_map = {
                'brand':              request.POST.getlist('brand[]'),
                'side':               request.POST.getlist('side[]'),
                'part_no':            request.POST.getlist('part_no[]'),
                'serial_no':          request.POST.getlist('serial_no[]'),
                'specs':              request.POST.getlist('specs[]'),
                'barcode':            request.POST.getlist('barcode[]'),
                'supported_model':    request.POST.getlist('supported_model[]'),
                'qty':                request.POST.getlist('qty[]'),
                'location':           request.POST.getlist('location[]'),
                'reference_location': request.POST.getlist('reference_location[]'),
                'remark':             request.POST.getlist('remark[]'),
                'store_location':     request.POST.getlist('store_location[]'),
                'stock_status':       request.POST.getlist('stock_status[]'),
            }
 
            total = len(data_map['serial_no'])
 
            for i in range(total):
                serial = data_map['serial_no'][i].strip()
                if not serial:
                    continue
 
                row = {}
                for key, value_list in data_map.items():
                    try:
                        row[key] = value_list[i]
                    except IndexError:
                        row[key] = None
 
                if not row.get('store_location'):
                    row['store_location'] = 'WH1'
                if not row.get('stock_status'):
                    row['stock_status'] = 'LIVE'
 
                product = create_product_and_railkit(row)
                _mark_latest_stock_in_user(product, request.user)
 
            messages.success(request, 'Rail Kit added successfully.')
            return redirect('add_railkit')
 
        except Exception as e:
            messages.error(request, str(e))
 
    return render(request, 'railkit/add_railkit.html', {
        'brands': Brand.objects.all(),
    })
 
 
def railkit_list(request):
    from apps.categories.models import RailKit
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    railkits = RailKit.objects.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date = Subquery(latest_audit.values('audited_on')[:1]),
    )
    railkits = _exclude_out_or_frozen(railkits)
 
    return render(request, 'railkit/railkit_list.html', {'railkits': railkits})
 
 
def update_railkit_field(request):
    from apps.categories.models import RailKit
 
    ALLOWED_FIELDS = ['remark']
 
    if request.method == 'POST':
        rk_id = request.POST.get('id')
        field = request.POST.get('field')
        value = request.POST.get('value')
 
        if field not in ALLOWED_FIELDS:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            rk = RailKit.objects.get(id=rk_id)
            setattr(rk, field, append_dated_remark(getattr(rk, field), value))
            rk.save()
            return JsonResponse({'success': True})
        except RailKit.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
 
    return JsonResponse({'success': False})


def add_harddisk(request):
    if request.method == 'POST':
        try:
            data_map = {
                'brand':              request.POST.getlist('brand[]'),
                'oem':                request.POST.getlist('oem[]'),
                'brand_model_no':     request.POST.getlist('brand_model_no[]'),
                'oem_model_no':       request.POST.getlist('oem_model_no[]'),
                'capacity':           request.POST.getlist('capacity[]'),
                'rpm':                request.POST.getlist('rpm[]'),
                'interface':          request.POST.getlist('interface[]'),
                'part_no':            request.POST.getlist('part_no[]'),
                'alt_part_no':        request.POST.getlist('alt_part_no[]'),
                'alt_fru_1':          request.POST.getlist('alt_fru_1[]'),
                'alt_fru_2':          request.POST.getlist('alt_fru_2[]'),
                'alt_fru_3':          request.POST.getlist('alt_fru_3[]'),
                'retail_part_no':     request.POST.getlist('retail_part_no[]'),
                'spare_part_tray':    request.POST.getlist('spare_part_tray[]'),
                'gpn_code':           request.POST.getlist('gpn_code[]'),
                'brand_serial_no':    request.POST.getlist('brand_serial_no[]'),
                'oem_serial_no':      request.POST.getlist('oem_serial_no[]'),
                'serial_no':          request.POST.getlist('serial_no[]'),
                'firmware':           request.POST.getlist('firmware[]'),
                'size':               request.POST.getlist('size[]'),
                'health':             request.POST.getlist('health[]'),
                'gb_s':               request.POST.getlist('gb_s[]'),
                'barcode':            request.POST.getlist('barcode[]'),
                'tray_barcode':       request.POST.getlist('tray_barcode[]'),
                'location':           request.POST.getlist('location[]'),
                'reference_location': request.POST.getlist('reference_location[]'),
                'remark':             request.POST.getlist('remark[]'),
                'store_location':     request.POST.getlist('store_location[]'),
                'stock_status':       request.POST.getlist('stock_status[]'),
            }
 
            total = len(data_map['serial_no'])
 
            for i in range(total):
                serial = data_map['serial_no'][i].strip()
                if not serial:
                    continue
 
                row = {}
                for key, value_list in data_map.items():
                    try:
                        row[key] = value_list[i]
                    except IndexError:
                        row[key] = None
 
                if not row.get('store_location'):
                    row['store_location'] = 'WH1'
                if not row.get('stock_status'):
                    row['stock_status'] = 'LIVE'
 
                product = create_product_and_harddisk(row)
                _mark_latest_stock_in_user(product, request.user)
 
            messages.success(request, 'Hard Disk added successfully.')
            return redirect('add_harddisk')
 
        except Exception as e:
            messages.error(request, str(e))
 
    return render(request, 'harddisk/add_harddisk.html', {
        'brands': Brand.objects.all(),
    })
 
 
def harddisk_list(request):
    from apps.categories.models import HardDisk
 
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
 
    latest_audit = InventoryTransaction.objects.filter(
        product=OuterRef('product'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
 
    harddisks = HardDisk.objects.select_related(
        'product', 'product__category', 'brand'
    ).annotate(
        latest_type     = Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location = Subquery(latest_txn.values('store_location')[:1]),
        latest_status   = Subquery(latest_txn.values('stock_status')[:1]),
        last_audit_date = Subquery(latest_audit.values('audited_on')[:1]),
    )
    harddisks = _exclude_out_or_frozen(harddisks)
 
    return render(request, 'harddisk/harddisk_list.html', {'harddisks': harddisks})
 
 
def update_harddisk_field(request):
    from apps.categories.models import HardDisk
 
    ALLOWED_FIELDS = ['remark']
 
    if request.method == 'POST':
        hd_id = request.POST.get('id')
        field = request.POST.get('field')
        value = request.POST.get('value')
 
        if field not in ALLOWED_FIELDS:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
 
        try:
            hd = HardDisk.objects.get(id=hd_id)
            setattr(hd, field, append_dated_remark(getattr(hd, field), value))
            hd.save()
            return JsonResponse({'success': True})
        except HardDisk.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
 
    return JsonResponse({'success': False})


def check_barcode(request):
    """
    AJAX endpoint — checks barcode uniqueness across ALL
    category tables in a single efficient query.
    """
    from apps.categories.models import (
        Spare, Card, CPU, Memory, SFP, RailKit, HardDisk, Controller
    )
 
    barcode = request.GET.get('barcode', '').strip().upper()
 
    if not barcode:
        return JsonResponse({'exists': False})
 
    exists = (
        Spare.objects.filter(barcode=barcode).exists()       or
        Card.objects.filter(barcode=barcode).exists()        or
        CPU.objects.filter(barcode=barcode).exists()         or
        Memory.objects.filter(barcode=barcode).exists()      or
        SFP.objects.filter(barcode=barcode).exists()         or
        RailKit.objects.filter(barcode=barcode).exists()     or
        HardDisk.objects.filter(barcode=barcode).exists()    or
        HardDisk.objects.filter(tray_barcode=barcode).exists() or
        Controller.objects.filter(barcode=barcode).exists()
    )
 
    return JsonResponse({'exists': exists})



def _annotate_membership(queryset, product_field='product'):
    """
    Annotates any category queryset with:
      - server_id, server_label  (if product is in a server)
      - controller_id, controller_label (if product is in a controller)
 
    product_field: the field path to the Product FK on the queryset model
    """
    from apps.servers.models import ServerComponent
    from apps.categories.models import Controller
 
    # Server membership
    server_comp = ServerComponent.objects.filter(
        product=OuterRef(product_field)
    ).select_related('server').order_by('-attached_at')
 
    # Controller membership (via Spare.controller)
    # For spare: OuterRef('id') → spare.controller_id
    # For other models: product → spare → controller (handled differently)
 
    queryset = queryset.annotate(
        server_label=Subquery(
            ServerComponent.objects.filter(
                product=OuterRef(product_field)
            ).values('server__machine_no')[:1]
        ),
        server_model_label=Subquery(
            ServerComponent.objects.filter(
                product=OuterRef(product_field)
            ).values('server__model')[:1]
        ),
        server_service_tag=Subquery(
            ServerComponent.objects.filter(
                product=OuterRef(product_field)
            ).values('server__service_tag')[:1]
        ),
        server_pk=Subquery(
            ServerComponent.objects.filter(
                product=OuterRef(product_field)
            ).values('server__id')[:1]
        ),
    )
 
    return queryset


from django.db.models import OuterRef, Subquery
from django.shortcuts import render

from apps.inventory.models import InventoryTransaction
from apps.servers.models import ServerComponent

from apps.categories.models import (
    Card,
    CPU,
    Memory,
    SFP,
    RailKit,
    HardDisk,
)


def _list_view_annotated(model_class, template, context_key):
    """
    Generic annotated list view for spare categories.
    """

    def view(request):

        latest_txn = InventoryTransaction.objects.filter(
            product_id=OuterRef('product_id')
        ).order_by('-created_at')

        latest_audit = InventoryTransaction.objects.filter(
            product_id=OuterRef('product_id'),
            transaction_type='AUDIT'
        ).order_by('-created_at')

        # IMPORTANT:
        # use product_id instead of product
        server_component = ServerComponent.objects.filter(
            product_id=OuterRef('product_id')
        )

        qs = (
            model_class.objects
            .select_related(
                'product',
                'product__category',
                'brand'
            )
            .annotate(

                # inventory
                latest_type=Subquery(
                    latest_txn.values('transaction_type')[:1]
                ),

                latest_location=Subquery(
                    latest_txn.values('store_location')[:1]
                ),

                latest_status=Subquery(
                    latest_txn.values('stock_status')[:1]
                ),

                last_audit_date=Subquery(
                    latest_audit.values('audited_on')[:1]
                ),

                # server membership
                server_label=Subquery(
                    server_component.values('server__machine_no')[:1]
                ),

                server_service_tag=Subquery(
                    server_component.values('server__service_tag')[:1]
                ),

                server_model_label=Subquery(
                    server_component.values('server__model')[:1]
                ),

                server_pk=Subquery(
                    server_component.values('server__id')[:1]
                ),
            )
        )
        qs = _exclude_out_or_frozen(qs)

        return render(
            request,
            template,
            {
                context_key: qs
            }
        )

    return view


# views
card_list = _list_view_annotated(
    Card,
    'card/list.html',
    'cards'
)

cpu_list = _list_view_annotated(
    CPU,
    'cpu/cpu_list.html',
    'cpus'
)

memory_list = _list_view_annotated(
    Memory,
    'memory/memory_list.html',
    'memories'
)

sfp_list = _list_view_annotated(
    SFP,
    'sfp/sfp_list.html',
    'sfps'
)

railkit_list = _list_view_annotated(
    RailKit,
    'railkit/railkit_list.html',
    'railkits'
)

harddisk_list = _list_view_annotated(
    HardDisk,
    'harddisk/harddisk_list.html',
    'harddisks'
)


def _latest_txn_subquery():
    return InventoryTransaction.objects.filter(
        product_id=OuterRef('product_id')
    ).order_by('-created_at')


def _annotated_category_queryset(model_class, sold=False, include_stocked_out=False):
    latest_txn = _latest_txn_subquery()
    latest_audit = InventoryTransaction.objects.filter(
        product_id=OuterRef('product_id'),
        transaction_type='AUDIT'
    ).order_by('-created_at')
    qs = model_class.objects.select_related('product', 'product__category', 'brand').annotate(
        latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
        latest_location=Subquery(latest_txn.values('store_location')[:1]),
        latest_status=Subquery(latest_txn.values('stock_status')[:1]),
        latest_client=Subquery(latest_txn.values('client_name')[:1]),
        latest_invoice=Subquery(latest_txn.values('invoice_no')[:1]),
        latest_olf_dc=Subquery(latest_txn.values('olf_dc_number')[:1]),
        latest_out_date=Subquery(latest_txn.values('stock_out_date')[:1]),
        last_audit_date=Subquery(latest_audit.values('audited_on')[:1]),
    )
    if sold:
        return qs.filter(latest_type='OUT')
    return qs if include_stocked_out else _exclude_out_or_frozen(qs)


LIST_MODELS = {
    'spare': {
        'label': 'Spare',
        'model': Spare,
        'add_url': 'add_spare',
        'list_url': 'spare_list',
        'sold_url': 'inventory_sold',
        'fields': [
            ('Product', 'product.name'),
            ('Category', 'product.category.name'),
            ('Brand', 'brand.name'),
            ('Model', 'model'),
            ('Part No', 'part_no'),
            ('Alt Part', 'alt_part_no'),
            ('Serial', 'product.serial_no'),
            ('Alt Serial', 'alt_serial_no'),
            ('Specs', 'specs'),
            ('Qty', 'qty'),
            ('Barcode', 'barcode'),
            ('Location', 'location'),
            ('Reference Location', 'reference_location'),
            ('Remark', 'remark'),
        ],
    },
    'card': {
        'label': 'Card',
        'model': Card,
        'add_url': 'add_card',
        'list_url': 'card_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('OEM', 'oem'), ('Model', 'brand_model_no'), ('Interface', 'interface'), ('Part No', 'part_no'), ('Serial', 'brand_serial_no_1'), ('Barcode', 'barcode'), ('Location', 'location'), ('Remark', 'remark')],
    },
    'cpu': {
        'label': 'CPU',
        'model': CPU,
        'add_url': 'add_cpu',
        'list_url': 'cpu_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('Model', 'model'), ('Part No', 'part_no'), ('Serial', 'product.serial_no'), ('Cores', 'no_of_cores'), ('Threads', 'no_of_threads'), ('GHz', 'ghz'), ('Frequency', 'frequency'), ('Cache', 'cache'), ('Barcode', 'barcode'), ('Location', 'location'), ('Remark', 'remark')],
    },
    'controller': {
        'label': 'Controller',
        'model': Controller,
        'add_url': 'add_controller',
        'list_url': 'controller_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('Model', 'model'), ('Part No', 'part_no'), ('Serial', 'product.serial_no'), ('Specs', 'specs'), ('Barcode', 'barcode'), ('Location', 'location'), ('Reference Location', 'reference_location'), ('Parent-child Location', 'parent_child_location'), ('Remark', 'remark')],
    },
    'memory': {
        'label': 'Memory',
        'model': Memory,
        'add_url': 'add_memory',
        'list_url': 'memory_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('OEM', 'oem'), ('Model', 'model'), ('Part No 1', 'part_no_1'), ('Part No 2', 'part_no_2'), ('Serial', 'product.serial_no'), ('Size', 'size'), ('RAM Type', 'ram_type'), ('Barcode', 'barcode'), ('Location', 'location'), ('Remark', 'remark')],
    },
    'sfp': {
        'label': 'SFP',
        'model': SFP,
        'add_url': 'add_sfp',
        'list_url': 'sfp_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('Model', 'model'), ('Part No', 'part_no'), ('Description', 'description'), ('Fibre Type', 'fibre_type'), ('Data Rate', 'data_rate'), ('Serial', 'product.serial_no'), ('Barcode', 'barcode'), ('Location', 'location'), ('Remark', 'remark')],
    },
    'railkit': {
        'label': 'Rail Kit',
        'model': RailKit,
        'add_url': 'add_railkit',
        'list_url': 'railkit_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('Side', 'side'), ('Part No', 'part_no'), ('Serial', 'product.serial_no'), ('Specs', 'specs'), ('Supported Model', 'supported_model'), ('Barcode', 'barcode'), ('Location', 'location'), ('Remark', 'remark')],
    },
    'harddisk': {
        'label': 'Hard Disk',
        'model': HardDisk,
        'add_url': 'add_harddisk',
        'list_url': 'harddisk_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('OEM', 'oem'), ('Brand Model', 'brand_model_no'), ('OEM Model', 'oem_model_no'), ('Capacity', 'capacity'), ('RPM', 'rpm'), ('Interface', 'interface'), ('Size', 'size'), ('Part No', 'part_no'), ('Serial', 'product.serial_no'), ('Barcode', 'barcode'), ('Tray Barcode', 'tray_barcode'), ('Location', 'location'), ('Remark', 'remark')],
    },
    'networking_spare': {
        'label': 'Networking Spare',
        'model': __import__('apps.categories.models', fromlist=['NetworkingSpare']).NetworkingSpare,
        'add_url': 'add_networking_spare',
        'list_url': 'networking_spare_list',
        'fields': [('Product', 'product.name'), ('Brand', 'brand.name'), ('Part No', 'part_no'), ('Alt Part', 'alt_part_no'), ('Serial', 'product.serial_no'), ('Alt Serial', 'alt_serial_no'), ('Specs', 'specs'), ('Qty', 'qty'), ('Barcode', 'barcode'), ('Location', 'location'), ('Reference Location', 'reference_location'), ('Remark', 'remark')],
    },
}


def _resolve_attr(obj, path):
    current = obj
    for part in path.split('.'):
        current = getattr(current, part, None)
        if current is None:
            return ''
    return current


def _generic_list_context(kind, sold=False, status=None, filters=None):
    config = LIST_MODELS[kind]
    filters = filters or {}
    base_items = _annotated_category_queryset(config['model'], sold=sold)
    total_count = base_items.count()
    items = base_items
    if sold and status:
        items = items.filter(latest_status=status)

    q = (filters.get('q') or '').strip()
    date_from = (filters.get('date_from') or '').strip()
    date_to = (filters.get('date_to') or '').strip()
    if sold and date_from:
        items = items.filter(latest_out_date__gte=date_from)
    if sold and date_to:
        items = items.filter(latest_out_date__lte=date_to)
    if q:
        q_obj = (
            Q(product__name__icontains=q) |
            Q(product__serial_no__icontains=q) |
            Q(product__part_no__icontains=q) |
            Q(barcode__icontains=q)
        )
        if sold:
            q_obj |= (
                Q(latest_client__icontains=q) |
                Q(latest_invoice__icontains=q) |
                Q(latest_olf_dc__icontains=q)
            )
        items = items.filter(q_obj)

    rows = []
    for item in items:
        rows.append({
            'item': item,
            'values': [_resolve_attr(item, path) for _, path in config['fields']],
        })
    return {
        'kind': kind,
        'title': ('Sold ' if sold else '') + config['label'],
        'config': config,
        'fields': config['fields'],
        'items': items,
        'rows': rows,
        'sold': sold,
        'total_count': total_count,
        'result_count': len(rows),
        'filters': {'q': q, 'date_from': date_from, 'date_to': date_to},
    }


@login_required
def inventory_sold_list(request, kind):
    if kind not in LIST_MODELS:
        return JsonResponse({'error': 'Unknown list'}, status=404)
    selected_status = request.GET.get('status', '').strip() or 'SALE'
    config = LIST_MODELS[kind]
    all_sold_qs = _annotated_category_queryset(config['model'], sold=True)
    available_statuses = sorted(list(set(
        all_sold_qs.values_list('latest_status', flat=True)
    )))
    available_statuses = [s for s in available_statuses if s]

    filters = {
        'q': request.GET.get('q', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    context = _generic_list_context(kind, sold=True, status=selected_status, filters=filters)
    context['available_statuses'] = available_statuses
    context['selected_status'] = selected_status
    context['can_stock_return'] = has_permission(request.user, 'stock_return')
    return render(request, 'inventory/generic_sold_list.html', context)


@login_required
@require_any_permission('sold_view', 'reports')
def inventory_faulty_list(request, kind):
    if kind not in LIST_MODELS:
        return JsonResponse({'error': 'Unknown list'}, status=404)
    config = LIST_MODELS[kind]
    items = _annotated_category_queryset(config['model'], include_stocked_out=True).filter(latest_status__in=('FAULTY', 'DAMAGED'))
    q = request.GET.get('q', '').strip()
    if q:
        items = items.filter(Q(product__name__icontains=q) | Q(product__serial_no__icontains=q) | Q(barcode__icontains=q))
    rows = [{'item': item, 'values': [_resolve_attr(item, path) for _, path in config['fields']]} for item in items]
    return render(request, 'inventory/faulty_list.html', {
        'kind': kind,
        'config': config,
        'fields': config['fields'],
        'rows': rows,
        'q': q,
        'can_stock_return': has_permission(request.user, 'stock_return'),
        'can_stock_out': has_permission(request.user, 'stock_out'),
    })


@login_required
def export_inventory(request, kind, state='live'):
    if kind not in LIST_MODELS:
        return JsonResponse({'error': 'Unknown export'}, status=404)
    import csv
    config = LIST_MODELS[kind]
    sold = state == 'sold'
    qs = _annotated_category_queryset(config['model'], sold=sold)
    if sold:
        selected_status = request.GET.get('status', '').strip()
        if selected_status:
            qs = qs.filter(latest_status=selected_status)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{kind}-{state}.csv"'
    writer = csv.writer(response)
    writer.writerow([label for label, _ in config['fields']] + [
        'Store', 'Status', 'Client', 'Invoice', 'OLF / DC No', 'Stock Out Date', 'Last Audit', 'Created On'
    ])
    for item in qs.iterator(chunk_size=1000):
        writer.writerow([
            _resolve_attr(item, path) for _, path in config['fields']
        ] + [
            item.latest_location or '',
            item.latest_status or '',
            item.latest_client or '',
            item.latest_invoice or '',
            item.latest_olf_dc or '',
            item.latest_out_date or '',
            item.last_audit_date or '',
            item.created_at,
        ])
    return response


CONTROLLER_EXPORT_HEADERS = [
    'Sr.No', 'Product Category', 'Product', 'Brand', 'Model', 'Part No',
    'Alt Part No', 'Serial No', 'Alt Serial. no', 'Specs', 'QTY', 'Barcode No',
    'LOCATION', 'Reference Location', 'Parent-child Location',
    'Remark(describe exact issue)',
]


@login_required
def export_controllers(request, state='live'):
    """Grouped XLSX export: each parent controller followed by its in-stock
    components, in the exact Controller import-template layout."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from apps.categories.models import Controller

    sold = state == 'sold'
    latest_txn = InventoryTransaction.objects.filter(
        product=OuterRef('product')
    ).order_by('-created_at')
    controllers = Controller.objects.select_related('product', 'brand').annotate(
        latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
    )
    controllers = (controllers.filter(latest_type='OUT') if sold
                   else _exclude_out_or_frozen(controllers)).order_by('id')

    def in_stock_components(controller):
        freeze = InventoryFreezeRecord.objects.filter(
            product=OuterRef('product')
        ).order_by('-frozen_at', '-id')
        comps = (controller.components
                 .select_related('product', 'product__category', 'brand')
                 .annotate(
                     latest_type=Subquery(latest_txn.values('transaction_type')[:1]),
                     freeze_status=Subquery(freeze.values('status')[:1]),
                 ).order_by('id'))
        if sold:
            return comps.filter(latest_type='OUT')
        return comps.exclude(latest_type='OUT').filter(
            Q(freeze_status__isnull=True) | ~Q(freeze_status='FROZEN')
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Controllers'
    ws.append(CONTROLLER_EXPORT_HEADERS)
    header_fill = PatternFill(fill_type='solid', fgColor='DCEBFF')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    group = 0
    for ctrl in controllers:
        group += 1
        parent_serial = ctrl.product.serial_no if ctrl.product else ''
        ws.append([
            group, 'CONTROLLER', ctrl.product.name if ctrl.product else 'CONTROLLER',
            ctrl.brand.name if ctrl.brand else '',
            ctrl.model or '', ctrl.part_no or '', ctrl.alt_part_no or '',
            parent_serial, ctrl.alt_serial_no or '', ctrl.specs or '',
            ctrl.qty or 1, ctrl.barcode or '', ctrl.location or '',
            ctrl.reference_location or '', '', ctrl.remark or '',
        ])
        for c in in_stock_components(ctrl):
            product_type = c.product.category.name if c.product and c.product.category else 'SPARE'
            ws.append([
                group, product_type, c.product.name if c.product else product_type,
                c.brand.name if c.brand else '',
                c.model or '', c.part_no or '', c.alt_part_no or '',
                c.product.serial_no if c.product else '', c.alt_serial_no or '',
                c.specs or '', c.qty or 1, c.barcode or '',
                c.location or ctrl.location or '', c.reference_location or '',
                parent_serial, c.remark or '',
            ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="controllers-{state}.xlsx"'
    return response


@login_required
def import_inventory_page(request):
    from apps.core.importers import IMPORT_LABELS
    stock_in_labels = {}
    stock_out_labels = {}
    for key, label in IMPORT_LABELS.items():
        if key.endswith('_stock_out') or key == 'stock_out':
            stock_out_labels[key] = label
        else:
            stock_in_labels[key] = label
    can_stock_in = has_permission(request.user, 'stock_in')
    can_stock_out = has_permission(request.user, 'stock_out_import')
    if not (can_stock_in or can_stock_out):
        return HttpResponseForbidden('Permission denied')
    from apps.inventory.models import InventoryTransaction
    return render(request, 'inventory/import.html', {
        'import_labels': IMPORT_LABELS,
        'stock_in_labels': stock_in_labels,
        'stock_out_labels': stock_out_labels,
        'can_stock_in': can_stock_in,
        'can_stock_out': can_stock_out,
        'store_locations': InventoryTransaction.STORE_LOCATION,
    })


def import_template_download(request, model_key):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    from apps.core.importers import HEADER_MAPS, IMPORT_LABELS, STOCK_OUT_APPEND_COLUMNS
    is_stock_out = model_key == 'stock_out_columns' or model_key == 'stock_out' or model_key.endswith('_stock_out')
    if not (has_permission(request.user, 'stock_out_import') if is_stock_out else has_permission(request.user, 'stock_in')):
        return HttpResponseForbidden('Permission denied')

    # Special template: ONLY the appended Stock Out columns, to paste at the
    # end of any existing Stock In Excel.
    if model_key == 'stock_out_columns':
        headers = list(STOCK_OUT_APPEND_COLUMNS.keys())
        label = 'Stock Out Columns'
    elif model_key not in HEADER_MAPS:
        return JsonResponse({'error': 'Unknown import template'}, status=404)
    else:
        label = IMPORT_LABELS.get(model_key, model_key)
        headers = list(HEADER_MAPS[model_key].keys())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = label[:31]
    sheet.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='DCEBFF')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 3, 14), 42)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{model_key}-import-template.xlsx"'
    return response


@login_required
@require_POST
def start_import(request):
    from apps.core.importers import start_import_job
    from apps.inventory.models import InventoryTransaction
    try:
        model_key = request.POST.get('model_key')
        is_stock_out = model_key == 'stock_out' or (model_key or '').endswith('_stock_out')
        if not (has_permission(request.user, 'stock_out_import') if is_stock_out else has_permission(request.user, 'stock_in')):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        upload = request.FILES.get('file')
        if not upload:
            return JsonResponse({'success': False, 'error': 'Please select an Excel file.'}, status=400)
        store_location = (request.POST.get('store_location') or '').strip().upper()
        valid_locations = dict(InventoryTransaction.STORE_LOCATION)
        if store_location and store_location not in valid_locations:
            return JsonResponse({'success': False, 'error': 'Please select a valid store location.'}, status=400)
        job = start_import_job(model_key, upload, request.user, store_location=store_location or None)
        return JsonResponse({'success': True, 'job_id': job.id, 'total_rows': job.total_rows})
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


@login_required
@require_POST
def process_import(request, job_id):
    from apps.core.importers import process_import_chunk
    from apps.categories.models import ImportJob
    try:
        job = ImportJob.objects.get(pk=job_id, created_by=request.user)
        is_stock_out = job.model_key == 'stock_out' or job.model_key.endswith('_stock_out')
        if not (has_permission(request.user, 'stock_out_import') if is_stock_out else has_permission(request.user, 'stock_in')):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        chunk_size = int(request.POST.get('chunk_size') or 100)
        job = process_import_chunk(job, chunk_size=chunk_size, user=request.user)
        return JsonResponse({
            'success': True,
            'status': job.status,
            'processed_rows': job.processed_rows,
            'total_rows': job.total_rows,
            'success_count': job.success_count,
            'error_count': job.error_count,
            'errors': job.errors[-10:],
            'percent': round((job.processed_rows / job.total_rows) * 100, 2) if job.total_rows else 100,
        })
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


def networking_spare_list(request):
    return render(request, 'inventory/generic_live_list.html', _generic_list_context('networking_spare', sold=False))


@login_required
def add_networking_spare(request):
    from apps.core.importers import create_product_and_networking_spare, mark_latest_stock_user
    if request.method == 'POST':
        try:
            total = len(request.POST.getlist('serial_no[]'))
            for i in range(total):
                row = {
                    'category': request.POST.getlist('category[]')[i] if i < len(request.POST.getlist('category[]')) else 'NETWORKING SPARE',
                    'brand': request.POST.getlist('brand[]')[i] if i < len(request.POST.getlist('brand[]')) else '',
                    'part_no': request.POST.getlist('part_no[]')[i] if i < len(request.POST.getlist('part_no[]')) else '',
                    'alt_part_no': request.POST.getlist('alt_part_no[]')[i] if i < len(request.POST.getlist('alt_part_no[]')) else '',
                    'serial_no': request.POST.getlist('serial_no[]')[i],
                    'alt_serial_no': request.POST.getlist('alt_serial_no[]')[i] if i < len(request.POST.getlist('alt_serial_no[]')) else '',
                    'specs': request.POST.getlist('specs[]')[i] if i < len(request.POST.getlist('specs[]')) else '',
                    'qty': request.POST.getlist('qty[]')[i] if i < len(request.POST.getlist('qty[]')) else 1,
                    'barcode': request.POST.getlist('barcode[]')[i] if i < len(request.POST.getlist('barcode[]')) else '',
                    'location': request.POST.getlist('location[]')[i] if i < len(request.POST.getlist('location[]')) else '',
                    'reference_location': request.POST.getlist('reference_location[]')[i] if i < len(request.POST.getlist('reference_location[]')) else '',
                    'remark': request.POST.getlist('remark[]')[i] if i < len(request.POST.getlist('remark[]')) else '',
                    'store_location': request.POST.getlist('store_location[]')[i] if i < len(request.POST.getlist('store_location[]')) else 'WH1',
                    'stock_status': request.POST.getlist('stock_status[]')[i] if i < len(request.POST.getlist('stock_status[]')) else 'LIVE',
                }
                if not row['serial_no'] and not row['barcode']:
                    continue
                product = create_product_and_networking_spare(row, user=request.user)
                mark_latest_stock_user(product, request.user)
            messages.success(request, 'Networking spare added successfully.')
            return redirect('networking_spare_list')
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, 'inventory/add_networking_spare.html')


def _product_membership(product_id):
    """Return server / controller membership info for a product.

    Used by universal search so a component is shown once (from its own
    category) with a highlighted "part of" label, rather than also listing
    the parent server / controller as a separate result.
    """
    from apps.servers.models import ServerComponent
    from apps.categories.models import Spare

    info = {'in_server': False, 'server_label': '', 'in_controller': False, 'controller_label': ''}

    sc = ServerComponent.objects.filter(product_id=product_id).select_related('server').first()
    if sc and sc.server:
        s = sc.server
        info['in_server'] = True
        info['server_label'] = (
            f"{s.machine_no or s.model or 'Server'} [{s.service_tag or '—'}]"
        )

    sp = Spare.objects.filter(
        product_id=product_id, controller__isnull=False
    ).select_related('controller__product').first()
    if sp and sp.controller:
        c = sp.controller
        info['in_controller'] = True
        info['controller_label'] = (
            f"{c.model or 'Controller'} [{getattr(c.product, 'serial_no', '') or '—'}]"
        )
    return info


def universal_search(request):
    query = (request.GET.get('q') or '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})

    results = []
    # Universal search covers every searchable field EXCEPT Product.name itself,
    # per spec — match identity, location, metadata and remarks.
    search_fields = ['product__serial_no', 'product__part_no', 'barcode',
                     'alt_part_no', 'alt_serial_no', 'location',
                     'reference_location', 'remark']
    from django.db.models import Q

    for kind, config in LIST_MODELS.items():
        q_obj = Q()
        model_fields = {f.name for f in config['model']._meta.get_fields()}
        for field in search_fields:
            local = field.split('__')[0]
            if local == 'product' or local in model_fields:
                q_obj |= Q(**{f'{field}__icontains': query})
        if not q_obj:
            continue
        for item in _annotated_category_queryset(config['model'], sold=False).filter(q_obj)[:8]:
            membership = _product_membership(item.product.id)
            results.append({
                'type': config['label'],
                'title': str(item.product.name),
                'serial': item.product.serial_no or '',
                'part_no': item.product.part_no or '',
                'barcode': getattr(item, 'barcode', '') or '',
                'location': getattr(item, 'location', '') or '',
                'status': item.latest_status or '',
                'product_id': item.product.id,
                'url': request.build_absolute_uri(),
                **membership,
            })

    # Servers are matched ONLY by their own identity. A component match surfaces
    # the component (above, from its category) with a "part of" label instead of
    # duplicating the parent server here.
    from apps.servers.models import Server
    server_q = (
        Q(machine_no__icontains=query) | Q(service_tag__icontains=query) |
        Q(model__icontains=query) | Q(part_no__icontains=query) |
        Q(barcode__icontains=query) | Q(location__icontains=query) |
        Q(alt_serial_no__icontains=query) | Q(alt_part_no__icontains=query)
    )
    for server in Server.objects.filter(server_q).distinct().select_related('product')[:10]:
        results.append({
            'type': 'Server',
            'title': server.model or 'Server',
            'serial': server.service_tag,
            'part_no': server.part_no or '',
            'barcode': server.barcode or '',
            'location': server.location or '',
            'status': server.status or '',
            'product_id': server.product.id if server.product else '',
            'url': request.build_absolute_uri('/servers/list/'),
            'in_server': False, 'server_label': '',
            'in_controller': False, 'controller_label': '',
        })

    return JsonResponse({'results': results[:50]})
