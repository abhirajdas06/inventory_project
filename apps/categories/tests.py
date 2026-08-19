from io import BytesIO
from datetime import date

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.categories.models import Card, Controller, ImportJob, Memory, NetworkingSpare, Spare
from apps.core.importers import import_controller_row
from apps.core.models import Brand, Product, SpareCategory, UserProfile
from apps.inventory.models import InventoryTransaction


class AuthAndNetworkingSpareTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='owner', password='pass12345')
        UserProfile.objects.create(user=self.user, role='ADMIN')

    def test_anonymous_can_view_live_list_but_cannot_add(self):
        live = self.client.get(reverse('networking_spare_list'))
        self.assertEqual(live.status_code, 200)

        add = self.client.get(reverse('add_networking_spare'))
        self.assertEqual(add.status_code, 302)
        self.assertIn(reverse('home'), add['Location'])

    def test_networking_spare_add_creates_product_and_stock_in_with_user(self):
        self.client.force_login(self.user)
        response = self.client.post(reverse('add_networking_spare'), {
            'category[]': ['WAN INTERFACE MODULE'],
            'brand[]': ['CISCO'],
            'part_no[]': ['73-9368-01'],
            'alt_part_no[]': ['800-24973-01'],
            'serial_no[]': ['FOC13233D4K'],
            'alt_serial_no[]': [''],
            'specs[]': ['SINGLE PORT ISDN'],
            'qty[]': ['1'],
            'barcode[]': ['MCMODA0300'],
            'location[]': ['Mwh2/Rack 60/Shelf 5/Box 5'],
            'reference_location[]': [''],
            'remark[]': [''],
            'store_location[]': ['WH1'],
            'stock_status[]': ['LIVE'],
        })

        self.assertEqual(response.status_code, 302)
        item = NetworkingSpare.objects.select_related('product').get(barcode='MCMODA0300')
        latest = InventoryTransaction.objects.filter(product=item.product).latest('created_at')
        self.assertEqual(latest.transaction_type, 'IN')
        self.assertEqual(latest.performed_by, self.user)

    def test_chunked_networking_spare_import(self):
        self.client.force_login(self.user)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'Product', 'Brand', 'Part no', 'Alt Part No', 'Serial No',
            'Alt Serial No', 'Specs', 'QTY', 'Barcode No', 'Location',
            'Reference Location', 'Remark (Describe exact Issue)'
        ])
        sheet.append([
            'WAN INTERFACE MODULE', 'CISCO', '73-9368-01', '800-24973-01',
            'FOC13233D4K', '', 'SINGLE PORT ISDN', 1, 'MCMODA0300',
            'Mwh2/Rack 60/Shelf 5/Box 5', '', ''
        ])
        sheet.append([
            'WAN INTERFACE MODULE', 'CISCO', '800-01514-02', 'WIC-1T',
            '37029907', '', 'SERIAL PORT', 1, 'MCMODA0301',
            'Mwh2/Rack 60/Shelf 5/Box 5', '', ''
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload = SimpleUploadedFile(
            'NETWORKING SPARE.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        start = self.client.post(reverse('inventory_import_start'), {
            'model_key': 'networking_spare',
            'file': upload,
        })
        self.assertEqual(start.status_code, 200)
        job_id = start.json()['job_id']

        process = self.client.post(reverse('inventory_import_process', args=[job_id]), {
            'chunk_size': 1,
        })
        self.assertEqual(process.status_code, 200)
        self.assertEqual(process.json()['processed_rows'], 1)
        self.assertEqual(process.json()['status'], 'RUNNING')

        process = self.client.post(reverse('inventory_import_process', args=[job_id]), {
            'chunk_size': 10,
        })
        self.assertEqual(process.status_code, 200)
        self.assertEqual(process.json()['status'], 'DONE')
        self.assertEqual(NetworkingSpare.objects.count(), 2)
        self.assertEqual(ImportJob.objects.get(pk=job_id).success_count, 2)

    def test_import_template_download_uses_expected_headers(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse('inventory_import_template', args=['networking_spare']))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        headers = [cell.value for cell in next(workbook.active.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers[:4], ['Product', 'Brand', 'Part no', 'Alt Part No'])
        self.assertIn('Barcode No', headers)

    def test_spare_remark_edit_appends_with_date(self):
        self.client.force_login(self.user)
        category = SpareCategory.objects.create(name='SPARE-REMARK')
        product = Product.objects.create(category=category, serial_no='RM-1', name='Remark test')
        spare = Spare.objects.create(product=product, barcode='RM-BC-1', remark='2026-08-18 - Initial note')

        response = self.client.post(reverse('update_spare'), {
            'id': spare.id,
            'field': 'remark',
            'value': 'Follow-up note',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        spare.refresh_from_db()
        self.assertIn('2026-08-18 - Initial note', spare.remark)
        self.assertIn('2026-08-19 - Follow-up note', spare.remark)

    def test_card_remark_edit_uses_shared_inline_endpoint(self):
        self.client.force_login(self.user)
        category = SpareCategory.objects.create(name='CARD-REMARK')
        product = Product.objects.create(category=category, serial_no='CARD-1', name='Card test')
        card = Card.objects.create(product=product, barcode='CARD-BC-1', remark='2026-08-18 - Existing card note')

        response = self.client.post(reverse('update_spare'), {
            'id': card.id,
            'model': 'card',
            'field': 'remark',
            'value': 'Fresh card note',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        card.refresh_from_db()
        self.assertIn('2026-08-18 - Existing card note', card.remark)
        self.assertIn('2026-08-19 - Fresh card note', card.remark)

    def test_memory_remark_edit_uses_memory_update_endpoint(self):
        self.client.force_login(self.user)
        category = SpareCategory.objects.create(name='MEMORY-REMARK')
        product = Product.objects.create(category=category, serial_no='MEM-1', name='Memory test')
        memory = Memory.objects.create(product=product, barcode='MEM-BC-1', remark='2026-08-18 - Existing memory note')

        response = self.client.post('/spare/memory-update/', {
            'id': memory.id,
            'field': 'remark',
            'value': 'Fresh memory note',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        memory.refresh_from_db()
        self.assertIn('2026-08-18 - Existing memory note', memory.remark)
        self.assertIn('2026-08-19 - Fresh memory note', memory.remark)

    def test_chunked_stock_out_import_by_barcode(self):
        self.client.force_login(self.user)
        category = SpareCategory.objects.create(name='SPARE')
        brand = Brand.objects.create(name='DELL')
        product = Product.objects.create(
            category=category,
            serial_no='STOCKOUT-001',
            name='Stock Out Test',
        )
        Spare.objects.create(
            product=product,
            brand=brand,
            part_no='PN-100',
            barcode='BC-STOCKOUT-001',
        )
        InventoryTransaction.objects.create(
            product=product,
            transaction_type='IN',
            store_location='WH1',
            stock_status='LIVE',
            performed_by=self.user,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'Serial No', 'Barcode No', 'Client Name', 'Invoice No',
            'OLF / DC No', 'Stock Status', 'Stock Out Date'
        ])
        sheet.append([
            '', 'BC-STOCKOUT-001', 'Import Client', 'INV-900',
            'OLF-900', 'SALE', '2026-06-26'
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            'STOCK_OUT.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        start = self.client.post(reverse('inventory_import_start'), {
            'model_key': 'stock_out',
            'file': upload,
        })
        self.assertEqual(start.status_code, 200)
        process = self.client.post(reverse('inventory_import_process', args=[start.json()['job_id']]), {
            'chunk_size': 10,
        })

        self.assertEqual(process.status_code, 200)
        self.assertEqual(process.json()['status'], 'DONE')
        latest = InventoryTransaction.objects.filter(product=product).latest('created_at')
        self.assertEqual(latest.transaction_type, 'OUT')
        self.assertEqual(latest.client_name, 'Import Client')
        self.assertEqual(latest.invoice_no, 'INV-900')
        self.assertEqual(latest.olf_dc_number, 'OLF-900')

    def test_stock_out_import_defaults_to_sale_and_today_when_optional_columns_missing(self):
        self.client.force_login(self.user)
        category = SpareCategory.objects.create(name='SPARE-DEFAULTS')
        product = Product.objects.create(
            category=category,
            serial_no='STOCKOUT-DEFAULT-001',
            name='Stock Out Default Test',
        )
        Spare.objects.create(
            product=product,
            barcode='BC-STOCKOUT-DEFAULT-001',
            part_no='PN-200',
        )
        InventoryTransaction.objects.create(
            product=product,
            transaction_type='IN',
            store_location='WH1',
            stock_status='LIVE',
            performed_by=self.user,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Serial No', 'Barcode No'])
        sheet.append(['', 'BC-STOCKOUT-DEFAULT-001'])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        upload = SimpleUploadedFile(
            'STOCK_OUT_DEFAULT.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        start = self.client.post(reverse('inventory_import_start'), {
            'model_key': 'stock_out',
            'file': upload,
        })
        self.assertEqual(start.status_code, 200)
        process = self.client.post(reverse('inventory_import_process', args=[start.json()['job_id']]), {
            'chunk_size': 10,
        })

        self.assertEqual(process.status_code, 200)
        self.assertEqual(process.json()['status'], 'DONE')
        latest = InventoryTransaction.objects.filter(product=product).latest('created_at')
        self.assertEqual(latest.transaction_type, 'OUT')
        self.assertEqual(latest.stock_status, 'SALE')
        self.assertEqual(latest.client_name, '')
        self.assertEqual(latest.invoice_no, '')
        self.assertEqual(latest.olf_dc_number, '')
        self.assertEqual(str(latest.stock_out_date), str(date.today()))

    def test_controller_import_creates_one_parent_controller_and_links_components(self):
        self.client.force_login(self.user)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'Sr.No', 'Product Category', 'Product', 'Brand', 'Model', 'Part No',
            'Alt Part No', 'Serial No', 'Alt Serial. no', 'Specs', 'QTY',
            'Barcode No', 'LOCATION', 'Reference Location', 'Parent-child Location',
            'Remark(describe exact issue)'
        ])
        sheet.append([
            4, 'CONTROLLER', '3 PAR 8450 CONTROLLER', 'HP', '3PAR STORESERV 8000', '792655-001',
            'H6Z17-63001', 'QEHCTA4D98Q0PS', '', '', 1, 'MCNTA001479',
            'Mwh1/Cupboard 18/Shelf 1/Box 1', 'CUB/18/P CABIN', '', 'T.OK SENTHIL (27/2/2024)'
        ])
        sheet.append([
            4, 'SFP', 'Short Wave Fibre Channel SFP+', 'HP', 'E7Y10A', '793444-001',
            '5697-3229', '7CR618J20R', '', '16GB SFF+SW XCVR-C', 1, 'MSFPA3991',
            'Mwh1/Cupboard 18/Shelf 1/Box 1', 'CUB/18/P CABIN', 'QEHCTA4D98Q0PS', ''
        ])
        sheet.append([
            4, 'BATTERY', 'Controller Battery Pack', 'HP', 'P9R41A', 'P9R41-63001',
            '', 'US0D668J757800931053', '',
            '', 1, 'MBTRA004320',
            'Mwh1/Cupboard 18/Shelf 1/Box 1', 'CUB/18/P CABIN',
            'QEHCTA4D98Q0PS', ''
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload = SimpleUploadedFile(
            'CONTROLLER.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        start = self.client.post(reverse('inventory_import_start'), {
            'model_key': 'controller',
            'file': upload,
        })
        self.assertEqual(start.status_code, 200)
        job_id = start.json()['job_id']

        first = self.client.post(reverse('inventory_import_process', args=[job_id]), {'chunk_size': 1})
        self.assertEqual(first.status_code, 200)
        self.assertEqual(Controller.objects.count(), 1)
        self.assertEqual(Spare.objects.count(), 0)

        second = self.client.post(reverse('inventory_import_process', args=[job_id]), {'chunk_size': 10})
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.json()['status'], 'DONE')

        controller = Controller.objects.get()
        self.assertEqual(controller.product.serial_no, 'QEHCTA4D98Q0PS')
        self.assertEqual(controller.product.category.name, 'CONTROLLER')
        self.assertEqual(controller.product.name, '3 PAR 8450 CONTROLLER')
        self.assertEqual(Spare.objects.filter(controller=controller).count(), 2)
        self.assertTrue(Spare.objects.filter(controller=controller, product__category__name='SFP').exists())
        self.assertTrue(Spare.objects.filter(controller=controller, product__category__name='BATTERY').exists())

    def test_controller_import_carries_parent_context_across_chunks_and_blank_parent_cells(self):
        self.client.force_login(self.user)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'Sr.No', 'Product Category', 'Product', 'Brand', 'Model', 'Part No',
            'Alt Part No', 'Serial No', 'Alt Serial. no', 'Specs', 'QTY',
            'Barcode No', 'LOCATION', 'Reference Location', 'Parent-child Location',
            'Remark(describe exact issue)'
        ])
        sheet.append([
            1, 'CONTROLLER', 'NETAPP FAS2650', 'NETAPP', '', '111-02505+A7',
            '110-00546+A7', '21707020939', '', '', 1, 'MCNTA001793',
            'Mwh1/Rack 27/Shelf 4/Box 90', '', '', ''
        ])
        sheet.append([
            1, 'HARD DISK', '', 'NETAPP', '', 'MTFDDAT128MBF-1AN1ZABYY',
            '', '16511516E1C8', '', '128GB MSATA SSD', 1, 'MHDDA045427',
            'Mwh1/Rack 27/Shelf 4/Box 90', '', '21707020939', ''
        ])
        sheet.append([
            1, 'CARD', 'NETAPP FAS2650', 'NETAPP', '', '111-02732+A1',
            '', '21710018400', '', '16GB FC', 1, 'MCRDA017998',
            'Mwh1/Rack 27/Shelf 4/Box 90', '', '', ''
        ])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload = SimpleUploadedFile(
            'CONTROLLER.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        start = self.client.post(reverse('inventory_import_start'), {
            'model_key': 'controller',
            'file': upload,
        })
        self.assertEqual(start.status_code, 200)
        job_id = start.json()['job_id']

        first = self.client.post(reverse('inventory_import_process', args=[job_id]), {'chunk_size': 2})
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.json()['status'], 'RUNNING')

        second = self.client.post(reverse('inventory_import_process', args=[job_id]), {'chunk_size': 10})
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.json()['status'], 'DONE')

        controller = Controller.objects.get(product__serial_no='21707020939')
        self.assertEqual(Spare.objects.filter(controller=controller).count(), 2)
        self.assertTrue(Spare.objects.filter(controller=controller, product__serial_no='21710018400').exists())

    def test_controller_component_import_can_match_parent_by_alt_serial_and_blank_component_serial(self):
        self.client.force_login(self.user)
        controller_category = SpareCategory.objects.create(name='CONTROLLER')
        processor_category = SpareCategory.objects.create(name='PROCESSOR')
        brand = Brand.objects.create(name='DELL')

        controller_product = Product.objects.create(
            category=controller_category,
            serial_no='CN0654Y9FCP007CJ005GA01',
            part_no='0645Y9',
            brand='DELL',
            model='EMC SCV3020',
            name='E18M001',
        )
        controller = Controller.objects.create(
            product=controller_product,
            brand=brand,
            model='EMC SCV3020',
            part_no='0645Y9',
            alt_part_no='0RN6X8',
            alt_serial_no='CN0RN6X8FCP007CC0041A01',
            barcode='MCNTA001762',
        )
        InventoryTransaction.objects.create(
            product=controller_product,
            transaction_type='IN',
            store_location='WH1',
            stock_status='LIVE',
        )

        imported = import_controller_row({
            'spare_type': 'PROCESSOR',
            'product_name': 'INTEL XEON 1.70 GHZ',
            'brand': '',
            'model': 'E5-2603 V4',
            'part_no': 'SR2P0',
            'alt_part_no': '',
            'serial_no': '',
            'alt_serial_no': '',
            'specs': '',
            'qty': 1,
            'barcode': 'MCPUA011648',
            'location': 'Mwh1/Rack 27/Shelf 1/Box 1',
            'reference_location': '',
            'parent_child_location': 'CN0RN6X8FCP007CC0041A01',
            'remark': '',
        }, user=self.user)

        self.assertIsNotNone(imported.pk)
        spare = Spare.objects.get(barcode='MCPUA011648')
        self.assertEqual(spare.controller, controller)
        self.assertEqual(spare.product.category, processor_category)

    def test_controller_import_repairs_old_wrong_component_as_cabinet_record(self):
        self.client.force_login(self.user)
        cabinet_category = SpareCategory.objects.create(name='CONTROLLER')
        controller_category = SpareCategory.objects.create(name='SFP')

        cabinet_product = Product.objects.create(
            category=cabinet_category,
            serial_no='CN0770D8137400AU002BA00',
            part_no='0770D8',
            brand='DELL',
            model='MD3200I',
            name='3 PAR 8450 CONTROLLER',
        )
        controller = Controller.objects.create(
            product=cabinet_product,
            model='MD3200I',
            part_no='0770D8',
        )
        InventoryTransaction.objects.create(
            product=cabinet_product,
            transaction_type='IN',
            store_location='WH1',
            stock_status='LIVE',
        )

        bad_component_product = Product.objects.create(
            category=cabinet_category,
            serial_no='CN0Y990H137400B300JOA00',
            part_no='0Y990H',
            brand='DELL',
            model='MD3200I',
            name='Wrong Spare',
        )
        bad_component_controller = Controller.objects.create(
            product=bad_component_product,
            model='MD3200I',
            part_no='0Y990H',
        )
        InventoryTransaction.objects.create(
            product=bad_component_product,
            transaction_type='IN',
            store_location='WH1',
            stock_status='LIVE',
        )

        product = import_controller_row({
            'spare_type': 'SFP',
            'product_name': 'Short Wave Fibre Channel SFP+',
            'brand': 'HP',
            'model': 'E7Y10A',
            'part_no': '793444-001',
            'alt_part_no': '',
            'serial_no': 'CN0Y990H137400B300JOA00',
            'alt_serial_no': '',
            'specs': '16GB SFF+SW XCVR-C',
            'qty': 1,
            'barcode': 'MCRDA017778',
            'location': 'Mwh1/Cupboard 18/Shelf 1/Box 1',
            'reference_location': 'CUB/18/P CABIN',
            'parent_child_location': 'CN0770D8137400AU002BA00',
            'remark': '',
        }, user=self.user)

        self.assertEqual(product.pk, bad_component_product.pk)
        self.assertFalse(Controller.objects.filter(pk=bad_component_controller.pk).exists())
        self.assertTrue(Spare.objects.filter(controller=controller, product=bad_component_product).exists())
        bad_component_product.refresh_from_db()
        self.assertEqual(bad_component_product.category.name, controller_category.name)

    def test_sold_list_status_filtering(self):
        self.client.force_login(self.user)
        # Create a card product
        card_cat = SpareCategory.objects.create(name='CARD')
        brand = Brand.objects.create(name='DELL')
        p1 = Product.objects.create(
            category=card_cat,
            serial_no='SERIALCARD1',
            brand='DELL',
            model='MD3200',
            name='Card 1',
        )
        c1 = Card.objects.create(
            product=p1,
            brand=brand,
            brand_serial_no_1='SERIALCARD1',
            brand_model_no='MD3200',
        )
        p2 = Product.objects.create(
            category=card_cat,
            serial_no='SERIALCARD2',
            brand='DELL',
            model='MD3200',
            name='Card 2',
        )
        c2 = Card.objects.create(
            product=p2,
            brand=brand,
            brand_serial_no_1='SERIALCARD2',
            brand_model_no='MD3200',
        )
        # Create transactions: OUT with different stock_statuses
        InventoryTransaction.objects.create(
            product=p1,
            transaction_type='OUT',
            store_location='WH1',
            stock_status='SALE',
        )
        InventoryTransaction.objects.create(
            product=p2,
            transaction_type='OUT',
            store_location='WH1',
            stock_status='RENT',
        )

        # 1. Fetch the generic sold list for card
        url = reverse('inventory_sold', args=['card'])
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, 'SERIALCARD1')
        self.assertNotContains(res, 'SERIALCARD2')

        # 2. Filter by status=SALE
        res_sale = self.client.get(url + '?status=SALE')
        self.assertEqual(res_sale.status_code, 200)
        self.assertContains(res_sale, 'SERIALCARD1')
        self.assertNotContains(res_sale, 'SERIALCARD2')

        # 3. Export filtered CSV
        export_url = reverse('inventory_export', kwargs={'kind': 'card', 'state': 'sold'})
        export_res = self.client.get(export_url + '?status=SALE')
        self.assertEqual(export_res.status_code, 200)
        content = export_res.content.decode('utf-8')
        self.assertIn('SERIALCARD1', content)
        self.assertNotIn('SERIALCARD2', content)
